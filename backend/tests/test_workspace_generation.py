"""Excel/document generation + generated-file download tests."""
import io
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from main import app
from api.middleware.auth import require_auth
from api.db.base import SessionLocal
from api.db.models import User
from api.db import crud
from api.utils import workspace_storage
from api.services.workspace_agent.xlsx_builder import build_workbook, build_csv_bytes, WorkbookValidationError
from api.services.workspace_agent import doc_builder

USER = {"id": "ws-gen-user", "username": "gen@example.com", "name": "Gen User"}


def _ensure_user():
    s = SessionLocal()
    try:
        if not s.get(User, USER["id"]):
            s.add(User(id=USER["id"], username=USER["username"], name=USER["name"]))
            s.commit()
    finally:
        s.close()


@pytest.fixture
def client():
    _ensure_user()
    app.dependency_overrides[require_auth] = lambda: USER
    return TestClient(app)


def test_build_workbook_is_valid_and_reopenable():
    data = build_workbook([
        {"name": "Inventory", "headers": ["File", "Type", "Size"], "rows": [["a.pdf", "pdf", 1024], ["b.docx", "docx", 2048]]},
    ])
    wb = load_workbook(io.BytesIO(data))
    assert "Inventory" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    ws = wb["Inventory"]
    assert ws.cell(row=1, column=1).value == "File"
    assert ws.cell(row=2, column=1).value == "a.pdf"


def test_build_workbook_adds_validation_sheet_on_warnings():
    data = build_workbook(
        [{"name": "Data", "headers": ["X"], "rows": [[1]]}],
        warnings=["file.pdf: scanned page skipped"],
    )
    wb = load_workbook(io.BytesIO(data))
    assert "Validation" in wb.sheetnames


def test_build_workbook_requires_at_least_one_sheet():
    with pytest.raises(ValueError):
        build_workbook([])


def test_build_csv_bytes():
    data = build_csv_bytes(["a", "b"], [[1, 2], [3, 4]])
    text = data.decode("utf-8-sig")
    assert "a,b" in text
    assert "1,2" in text


def test_build_word_document_roundtrip():
    data = doc_builder.build_word_document("Report Title", "## Section\nSome body text.")
    from docx import Document
    doc = Document(io.BytesIO(data))
    all_text = "\n".join(p.text for p in doc.paragraphs)
    assert "Report Title" in all_text
    assert "Some body text" in all_text


def test_build_markdown_and_txt_reports():
    md = doc_builder.build_markdown_report("Title", "Body text").decode("utf-8")
    assert md.startswith("# Title")
    txt = doc_builder.build_txt_report("Title", "Body text").decode("utf-8")
    assert "Title" in txt and "Body text" in txt


def test_generated_file_download_headers(client):
    ws = crud.create_workspace(SessionLocal(), USER["id"], name="Download Test")
    data = build_workbook([{"name": "Sheet1", "headers": ["A"], "rows": [[1]]}])

    db = SessionLocal()
    try:
        storage_path = workspace_storage.save_generated(USER["id"], ws.id, "Inventory.xlsx", data)
        gf = crud.add_generated_file(
            db, ws.id, filename="Inventory.xlsx", format="xlsx",
            storage_path=storage_path, size_bytes=len(data),
        )
    finally:
        db.close()

    resp = client.get(f"/api/workspaces/{ws.id}/generated/{gf.id}/download")
    assert resp.status_code == 200
    assert "Inventory.xlsx" in resp.headers["content-disposition"]
    assert resp.headers["content-disposition"].startswith("attachment")
    # Confirm the downloaded bytes are a valid, reopenable workbook.
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Sheet1" in wb.sheetnames
