"""Workspace agent tool-calling loop tests, using a stub Gemini-shaped provider
(no real API key/network needed — exercises the loop's control flow, tool
dispatch, and task/event persistence)."""
import asyncio
import io
import json
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest

from api.db.base import SessionLocal
from api.db.models import User
from api.db import crud
from api.utils import workspace_storage
from api.services.workspace_agent.agent import run_workspace_turn


class _StubFn:
    def __init__(self, name, arguments):
        self.name = name
        self.arguments = arguments


class _StubToolCall:
    def __init__(self, call_id, name, arguments):
        self.id = call_id
        self.function = _StubFn(name, json.dumps(arguments))


class _StubMessage:
    def __init__(self, content=None, tool_calls=None):
        self.content = content
        self.tool_calls = tool_calls


class _StubProvider:
    provider_name = "Stub"
    model = "gemini-stub"

    def __init__(self, responses):
        self._responses = list(responses)
        self.calls = 0

    async def chat_with_tools(self, messages, tools, system_prompt=""):
        resp = self._responses[min(self.calls, len(self._responses) - 1)]
        self.calls += 1
        return resp

    async def chat(self, messages, system_prompt=""):
        return "stub reply"


AGENT_USER = {"id": "ws-agent-user", "username": "agent@example.com", "name": "Agent User"}


def _ensure_user():
    s = SessionLocal()
    try:
        if not s.get(User, AGENT_USER["id"]):
            s.add(User(id=AGENT_USER["id"], username=AGENT_USER["username"], name=AGENT_USER["name"]))
            s.commit()
    finally:
        s.close()


def _make_workspace_with_file():
    _ensure_user()
    db = SessionLocal()
    ws = crud.create_workspace(db, AGENT_USER["id"], name="Agent Test Workspace")
    data = b"The quick brown fox jumps over the lazy dog."
    storage_path = workspace_storage.save_upload(AGENT_USER["id"], ws.id, "notes.txt", data)
    wf = crud.add_workspace_file(
        db, ws.id, relative_path="notes.txt", original_filename="notes.txt",
        mime_type="text/plain", extension="txt", size_bytes=len(data),
        checksum=workspace_storage.checksum(data), storage_path=storage_path,
    )
    crud.update_workspace_file_status(db, wf.id, parse_status="ready")
    workspace_storage.save_extracted_text(AGENT_USER["id"], ws.id, wf.id, data.decode())
    crud.recompute_workspace_totals(db, ws.id)
    ws_id = ws.id
    db.close()
    return ws_id


class _FakeConversation:
    def __init__(self, conv_id):
        self.id = conv_id


async def _run(ws_id, message, provider):
    # Fresh session per call, matching how a real request-scoped `db` dependency
    # behaves — an ORM object must never be reused across sessions.
    db = SessionLocal()
    ws = crud.get_workspace(db, ws_id, AGENT_USER["id"])
    conv = _FakeConversation("fake-conv-id")
    events = []
    async for event in run_workspace_turn(db, ws, conv, message, [], provider, "gemini-stub"):
        events.append(event)
    db.close()
    return events


def test_agent_direct_answer_no_tool_calls():
    ws_id = _make_workspace_with_file()
    provider = _StubProvider([_StubMessage(content="Here is a summary of your workspace.")])
    events = asyncio.run(_run(ws_id, "Summarize this folder", provider))

    stages = [e["stage"] for e in events if e["type"] == "stage"]
    assert "understanding" in stages
    assert "completed" in stages

    done = next(e for e in events if e["type"] == "done")
    assert done["status"] == "completed"
    assert "summary of your workspace" in done["content"]


def test_agent_executes_tool_call_then_answers():
    ws_id = _make_workspace_with_file()
    tool_call = _StubToolCall("call_1", "list_workspace_files", {})
    provider = _StubProvider([
        _StubMessage(content=None, tool_calls=[tool_call]),
        _StubMessage(content="I found notes.txt in your workspace."),
    ])
    events = asyncio.run(_run(ws_id, "What files are in here?", provider))

    tool_calls = [e for e in events if e["type"] == "tool_call"]
    tool_results = [e for e in events if e["type"] == "tool_result"]
    assert len(tool_calls) == 1
    assert tool_calls[0]["tool"] == "list_workspace_files"
    assert tool_results[0]["ok"] is True

    done = next(e for e in events if e["type"] == "done")
    assert "notes.txt" in done["content"]

    # Task + events were persisted for audit/history.
    db = SessionLocal()
    task = crud.get_task(db, done["task_id"], ws_id)
    assert task is not None
    assert task.status == "completed"
    logged = crud.list_task_events(db, task.id)
    assert any(e.event_type == "tool_call" for e in logged)
    db.close()


def test_agent_generates_excel_and_registers_generated_file():
    ws_id = _make_workspace_with_file()
    args = {
        "filename": "inventory.xlsx",
        "sheets": [{"name": "Files", "headers": ["Name"], "rows": [["notes.txt"]]}],
    }
    tool_call = _StubToolCall("call_1", "create_excel_workbook", args)
    provider = _StubProvider([
        _StubMessage(content=None, tool_calls=[tool_call]),
        _StubMessage(content="Created an Excel inventory of your workspace."),
    ])
    events = asyncio.run(_run(ws_id, "Create an Excel inventory", provider))
    done = next(e for e in events if e["type"] == "done")
    assert len(done["generated_files"]) == 1
    assert done["generated_files"][0]["filename"] == "inventory.xlsx"

    db = SessionLocal()
    gen = crud.list_generated_files(db, ws_id)
    assert len(gen) == 1
    data = workspace_storage.read_file(gen[0].storage_path)
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    assert "Files" in wb.sheetnames
    db.close()


def test_agent_provider_without_tool_support_errors_cleanly():
    ws_id = _make_workspace_with_file()

    class _NoToolsProvider:
        provider_name = "NoTools"

        async def chat(self, messages, system_prompt=""):
            return "n/a"

    events = asyncio.run(_run(ws_id, "Summarize", _NoToolsProvider()))
    assert events[0]["type"] == "error"
    assert "tool calling" in events[0]["error"] or "does not support" in events[0]["error"]


def test_followup_turn_reuses_same_workspace_without_reupload():
    ws_id = _make_workspace_with_file()
    provider = _StubProvider([_StubMessage(content="First turn answer.")])
    events1 = asyncio.run(_run(ws_id, "Analyze this folder", provider))
    done1 = next(e for e in events1 if e["type"] == "done")

    provider2 = _StubProvider([_StubMessage(content="Second turn answer, same workspace.")])
    events2 = asyncio.run(_run(ws_id, "Now put the results into Excel", provider2))
    done2 = next(e for e in events2 if e["type"] == "done")

    assert done1["task_id"] != done2["task_id"]

    db = SessionLocal()
    files_after_both_turns = crud.list_workspace_files(db, ws_id)
    assert len(files_after_both_turns) == 1  # no re-upload happened, same single file throughout
    db.close()
