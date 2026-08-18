"""Professional Excel workbook generation for the workspace agent.

Real workbook generation via openpyxl (never a CSV-renamed-to-.xlsx). Every
workbook is reopened with ``openpyxl.load_workbook`` immediately after being
built to verify it is valid before the caller ever reports it as generated.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


class WorkbookValidationError(RuntimeError):
    """Raised when a freshly-built workbook fails to reopen for validation."""


def _style_header_row(ws) -> None:
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")


def _autosize(ws, num_cols: int, sample_rows: list[list]) -> None:
    widths = [0] * num_cols
    for row in sample_rows[:500]:
        for i, v in enumerate(row[:num_cols]):
            widths[i] = max(widths[i], len(str(v)) if v is not None else 0)
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = min(max(w + 2, 10), 60)


def _clean_cell(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date, int, float, str, bool)):
        return v
    return str(v)


def _write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title=(name or "Sheet")[:31])
    ws.append(headers)
    _style_header_row(ws)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([_clean_cell(v) for v in row])

    if headers:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autosize(ws, len(headers), [headers] + rows)
    return ws


def _unique_sheet_name(name: str, used: set[str]) -> str:
    base = (name or "Sheet")[:31]
    candidate = base
    n = 1
    while candidate in used:
        n += 1
        suffix = f" ({n})"
        candidate = base[: 31 - len(suffix)] + suffix
    used.add(candidate)
    return candidate


def build_workbook(
    sheets: list[dict],
    *,
    summary_title: str = "Summary",
    include_summary: bool = True,
    errors: list[str] | None = None,
    warnings: list[str] | None = None,
) -> bytes:
    """Build a multi-sheet .xlsx workbook.

    ``sheets``: [{"name": str, "headers": [str, ...], "rows": [[...], ...]}, ...]

    Adds a "Summary" sheet (sheet name / row count / column count) and, when
    errors/warnings are supplied, a "Validation" sheet — matching the spec's
    "summary worksheet" and "validation/error worksheet" requirements.

    Raises WorkbookValidationError if the generated file cannot be reopened.
    """
    if not sheets:
        raise ValueError("build_workbook requires at least one sheet")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    if include_summary:
        summary_rows = [[s["name"], len(s.get("rows", [])), len(s.get("headers", []))] for s in sheets]
        summary_ws = _write_sheet(wb, summary_title, ["Sheet", "Rows", "Columns"], summary_rows)
        wb.move_sheet(summary_ws, offset=-(len(wb.sheetnames) - 1))

    used_names = {ws_.title for ws_ in wb.worksheets}
    for s in sheets:
        name = _unique_sheet_name(s["name"], used_names)
        _write_sheet(wb, name, s.get("headers", []), s.get("rows", []))

    if warnings or errors:
        issue_rows = [["Error", e] for e in (errors or [])] + [["Warning", w] for w in (warnings or [])]
        _write_sheet(wb, "Validation", ["Type", "Message"], issue_rows)

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    try:
        verify_wb = load_workbook(io.BytesIO(data))
        verify_wb.close()
    except Exception as exc:
        raise WorkbookValidationError(f"Generated workbook failed to reopen for validation: {exc}") from exc

    return data


def build_csv_bytes(headers: list[str], rows: list[list]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows([[_clean_cell(v) for v in row] for row in rows])
    return buf.getvalue().encode("utf-8-sig")
