"""Excel (.xlsx/.xls) processor — openpyxl, per-sheet grid + first-row-header table."""
from __future__ import annotations

import io

from .base import make_result

_MAX_ROWS_READ = 2000
_MAX_PREVIEW_ROWS = 50
_MAX_TABLE_ROWS = 200


def process_excel(file_path: str, data: bytes) -> dict:
    warnings: list[str] = []
    errors: list[str] = []
    sheets_out: list[dict] = []
    tables: list[dict] = []

    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            rows: list[list] = []
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if v is None else v for v in row])
                if len(rows) >= _MAX_ROWS_READ:
                    warnings.append(f"Sheet '{name}' truncated to the first {_MAX_ROWS_READ} rows for processing.")
                    break

            sheets_out.append({
                "name": name,
                "row_count": ws.max_row or 0,
                "col_count": ws.max_column or 0,
                "preview_rows": rows[:_MAX_PREVIEW_ROWS],
            })
            if rows:
                header, *data_rows = rows
                tables.append({"sheet": name, "header": header, "rows": data_rows[:_MAX_TABLE_ROWS]})
        wb.close()

        if not sheets_out:
            warnings.append("Workbook has no sheets.")
    except Exception as exc:
        errors.append(f"Failed to read Excel workbook: {exc}")

    text_preview = "\n".join(
        f"[Sheet: {s['name']}] {s['row_count']} rows x {s['col_count']} cols" for s in sheets_out
    )
    return make_result(
        file_path, "xlsx",
        text=text_preview,
        sheets=sheets_out,
        tables=tables,
        metadata={"sheet_count": len(sheets_out)},
        warnings=warnings, errors=errors,
    )
