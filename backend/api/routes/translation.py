"""
Professional Technical Translation Studio — API routes.

Endpoints:
  POST   /api/translation/projects              — create project + upload document
  GET    /api/translation/projects              — list all projects
  GET    /api/translation/projects/{id}         — get project detail
  PATCH  /api/translation/projects/{id}         — rename, tags, settings
  DELETE /api/translation/projects/{id}         — delete project
  POST   /api/translation/projects/{id}/translate   — SSE translation pipeline
  POST   /api/translation/projects/{id}/versions    — save version snapshot
  POST   /api/translation/projects/{id}/versions/{vn}/restore — restore version
  PATCH  /api/translation/projects/{id}/segments/{sid}  — edit one segment
  GET    /api/translation/projects/{id}/export/{fmt}    — download file

  GET    /api/translation/dictionary            — list dictionary entries
  POST   /api/translation/dictionary            — add entry
  PUT    /api/translation/dictionary/{eid}      — update entry
  DELETE /api/translation/dictionary/{eid}      — delete entry
  POST   /api/translation/dictionary/import     — import TSV
  GET    /api/translation/dictionary/export     — export TSV

  GET    /api/translation/memory                — list memory entries
  PUT    /api/translation/memory/{mid}          — edit memory entry
  DELETE /api/translation/memory/{mid}          — delete memory entry

  GET    /api/translation/settings              — get user translation settings
  PUT    /api/translation/settings              — update user translation settings
"""
from __future__ import annotations

import asyncio
import io
import json
import logging
import os
import re
import shutil
import tempfile
import time
import uuid
import zipfile
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from sqlalchemy.orm.attributes import flag_modified

from api.db import get_db
from api.db.models import TranslationProject, TranslationSegment, CustomDictionaryEntry, ProviderConfig, TranslationUsage
from api.utils import cost_guard
from api.utils.filename_helper import content_disposition, mime_for_ext
from api.utils.file_storage import get_source_bytes, save_source_file, copy_source_file, delete_source_file, has_source_file
from api.middleware.auth import require_auth, require_admin_session
from api.providers.registry import provider_registry, encrypt_key, decrypt_key

log = logging.getLogger(__name__)
router = APIRouter(tags=["translation"])


def _native_office_enabled() -> bool:
    """Whether to use the native MS Office desktop COM 'translate-in-place' path.

    Translation Studio clone: default OFF. With it off, DOCX/PPTX always go through
    the reconstructed python-docx / python-pptx rebuild (doc_rebuilder), which
    carries the platform's Arabic RTL / bracket / image-layout work — AND matches
    the Linux deployment target, where MS Office COM is never available. The old
    native path (used only because PowerPoint/Word happen to be installed on this
    Windows box) bypassed rebuild_pptx/rebuild_docx and lost those fixes.
    Set NATIVE_OFFICE_TRANSLATE=true to opt back in.
    """
    return (os.environ.get("NATIVE_OFFICE_TRANSLATE", "false").strip().lower()
            in ("1", "true", "yes", "on"))


def _check_upload_disk_space(max_file_bytes: int) -> None:
    """
    Raise HTTP 507 if there is not enough free space in /tmp for the upload.

    Reserves 4× the maximum file size to cover:
      • the original uploaded file (1×)
      • extracted / working files during translation (1×)
      • the rebuilt translated output (1×)
      • a safety buffer (1×)
    """
    SAFETY_MULTIPLIER = 4
    required = max_file_bytes * SAFETY_MULTIPLIER
    try:
        usage = shutil.disk_usage("/tmp")
        if usage.free < required:
            free_mb = usage.free  // (1024 * 1024)
            need_mb = required    // (1024 * 1024)
            raise HTTPException(
                507,
                f"Insufficient disk space — {free_mb} MB free but {need_mb} MB required. "
                f"Please try again later or contact your administrator.",
            )
    except HTTPException:
        raise
    except Exception:
        pass  # Cannot stat /tmp — don't block the upload

# ── One-time glossary seed flag ───────────────────────────────────────────────
# seed_glossary is idempotent but runs 4 DB queries every call to check for
# existing entries. Skip it after the first successful seed in this process.
_GLOSSARY_SEEDED: bool = False

# ── Ownership helpers ──────────────────────────────────────────────────────────

def _user_id(user: Optional[dict]) -> Optional[str]:
    """Extract user_id from the auth dict (or None for anonymous)."""
    return user["id"] if user else None


def _get_owned_project(
    db: Session,
    project_id: str,
    user: Optional[dict],
) -> "TranslationProject":
    """
    Return the project if and only if the caller owns it.
    Returns HTTP 404 (not 403) to avoid leaking resource existence.
    Scopes by user_id (NULL == anonymous session).
    """
    uid = _user_id(user)
    p = (
        db.query(TranslationProject)
        .filter(
            TranslationProject.id == project_id,
            TranslationProject.user_id == uid,
        )
        .first()
    )
    if p is None:
        raise HTTPException(404, "Project not found")
    return p


def _get_owned_dict_entry(
    db: Session,
    entry_id: str,
    user: Optional[dict],
) -> "CustomDictionaryEntry":
    """Return a dictionary entry the caller owns (personal entries only, not shared)."""
    uid = _user_id(user)
    entry = (
        db.query(CustomDictionaryEntry)
        .filter(
            CustomDictionaryEntry.id == entry_id,
            CustomDictionaryEntry.user_id == uid,
        )
        .first()
    )
    if entry is None:
        raise HTTPException(404, "Entry not found")
    return entry


def _get_owned_memory_entry(
    db: Session,
    memory_id: str,
    user: Optional[dict],
) -> "TranslationSegment":
    """Return a memory entry the caller owns (personal entries only)."""
    uid = _user_id(user)
    entry = (
        db.query(TranslationSegment)
        .filter(
            TranslationSegment.id == memory_id,
            TranslationSegment.user_id == uid,
        )
        .first()
    )
    if entry is None:
        raise HTTPException(404, "Memory entry not found")
    return entry


# ── Misc helpers ───────────────────────────────────────────────────────────────

def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _sse(obj: dict) -> str:
    return f"data: {json.dumps(obj, default=str)}\n\n"


def _word_finalize_or_503(content: bytes, target_lang: str) -> bytes:
    """
    Mandatory final layout pass for a DOCX about to be delivered to the user.

    Cross-platform engine selection (Translation Studio clone):
      1. Microsoft Word desktop COM — used when running on Windows with Word
         available (highest fidelity; unchanged from the main platform).
      2. LibreOffice headless — the Linux deployment authority; used when Word
         COM is unavailable but LibreOffice is installed.
      3. Neither available → HTTP 503 with a clear, actionable message.

    No silent unfinalized bytes: if the selected engine fails, that error is
    surfaced rather than returning a non-finalized document.
    """
    from api.utils.document_finalizer import finalize_docx, DocumentFinalizeError

    try:
        return finalize_docx(content, target_lang)
    except DocumentFinalizeError as exc:
        log.exception("DOCX finalization failed during export")
        raise HTTPException(503, str(exc)) from exc


def _project_summary(p: TranslationProject) -> dict:
    return {
        "id": p.id,
        "name": p.name,
        "source_filename": p.source_filename,
        "source_file_type": p.source_file_type,
        "source_lang": p.source_lang,
        "target_lang": p.target_lang,
        "style": p.style,
        "status": p.status,
        "quality_score": p.quality_score,
        "quality_breakdown": p.quality_breakdown or {},
        "formatting_fidelity": getattr(p, "formatting_fidelity", None) or "reconstructed",
        "provider_name": p.provider_name or "auto",
        "segment_count": len(p.segments) if p.segments else 0,
        "version_num": p.version_num,
        "tags": p.tags or [],
        "created_at": p.created_at.isoformat() if p.created_at else None,
        "updated_at": p.updated_at.isoformat() if p.updated_at else None,
    }


def _project_detail(p: TranslationProject) -> dict:
    d = _project_summary(p)
    d["segments"] = p.segments or []
    d["quality_issues"] = p.quality_issues or []
    d["versions"] = p.versions or []
    d["keep_english_terms"] = p.keep_english_terms
    d["transliterate_names"] = p.transliterate_names
    d["engineering_review_changes"] = p.engineering_review_changes or []
    d["dnt_tokens"] = p.dnt_tokens or []
    return d


ALLOWED_EXTENSIONS = {
    "pdf", "docx", "pptx", "xlsx", "txt", "md", "html", "htm",
    "csv", "rtf", "xml", "odt",
}

MIME_TYPE_MAP = {
    "application/pdf": "pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "application/vnd.openxmlformats-officedocument.presentationml.presentation": "pptx",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "text/plain": "txt",
    "text/html": "html",
    "text/csv": "csv",
    "text/markdown": "md",
    "application/rtf": "rtf",
    "text/xml": "xml",
    "application/xml": "xml",
}

MAX_FILE_SIZE = 200 * 1024 * 1024  # 200 MB


def _get_file_type(filename: str, content_type: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ALLOWED_EXTENSIONS:
        return ext
    # Fall back to MIME
    ft = MIME_TYPE_MAP.get(content_type, "txt")
    return ft


# ── Project endpoints ─────────────────────────────────────────────────────────

@router.post("/translation/projects")
async def create_project(
    request: Request,
    file: UploadFile = File(...),
    name: str = Form(""),
    source_lang: str = Form("en"),
    target_lang: str = Form("ar"),
    style: str = Form("technical"),
    keep_english_terms: bool = Form(False),
    transliterate_names: bool = Form(True),
    provider_name: str = Form("auto"),
    layout_mode: str = Form("original"),
    style_profile_id: str = Form(""),
    template_strength: str = Form("balanced"),
    layout_options: str = Form("{}"),
    image_ocr_enabled: bool = Form(True),
    reference_template: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Upload a document and create a new translation project."""
    cost_guard.ensure_enabled()

    project_id = str(uuid.uuid4())
    max_bytes = cost_guard.max_file_size_bytes()

    # ── Disk-space preflight ───────────────────────────────────────────────────
    # Check before we accept a single byte — returns HTTP 507 with a clear
    # message if there is not enough free space in /tmp.
    _check_upload_disk_space(max_bytes)

    # ── Stream upload to a temp file (avoids loading 100 MB into event loop) ──
    # FastAPI's UploadFile.read(size) reads in chunks from python-multipart's
    # SpooledTemporaryFile.  We stream those chunks to our own temp file on disk
    # and enforce the size limit incrementally, so:
    #   • We never allocate a single large bytearray in the async event loop.
    #   • Oversized files are rejected mid-stream, not after full receipt.
    #   • Other coroutines can run between chunk reads (each await yields).
    tmp_path: Optional[str] = None
    file_bytes: bytes = b""
    received = 0
    try:
        tmp_fd, tmp_path = tempfile.mkstemp(prefix="translation_upload_", suffix=".tmp")
        CHUNK = 1 * 1024 * 1024  # 1 MB — small enough to yield often
        with os.fdopen(tmp_fd, "wb") as tmp_file:
            while True:
                chunk = await file.read(CHUNK)
                if not chunk:
                    break
                received += len(chunk)
                if received > max_bytes:
                    log.warning("Upload rejected: file too large for user %s (%d bytes > %d limit)", user.get("id"), received, max_bytes)
                    raise HTTPException(
                        400,
                        f"File too large — received more than "
                        f"{max_bytes // (1024 * 1024)} MB "
                        f"(limit configurable via MAX_FILE_SIZE_MB).",
                    )
                tmp_file.write(chunk)

        if received == 0:
            log.warning("Upload rejected: empty file for user %s", user.get("id"))
            raise HTTPException(400, "Empty file — no data was received.")

        # Read from disk in a thread (avoids blocking the event loop on I/O)
        def _read_tmp() -> bytes:
            with open(tmp_path, "rb") as fh:
                return fh.read()

        file_bytes = await asyncio.to_thread(_read_tmp)

    except HTTPException:
        raise
    except Exception as exc:
        log.exception("Upload storage failure for user %s: %s", user.get("id"), exc)
        raise HTTPException(500, f"Storage failure while saving upload: {exc}") from exc
    finally:
        # Always remove the temp file, whether the upload succeeded or failed
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    # Strict type enforcement: reject anything whose extension is not on the
    # allow-list AND whose declared MIME type is not recognized, instead of
    # silently treating it as plain text (which let disallowed files through).
    _raw_ext = (file.filename or "").rsplit(".", 1)[-1].lower() if "." in (file.filename or "") else ""
    if _raw_ext not in ALLOWED_EXTENSIONS and (file.content_type or "") not in MIME_TYPE_MAP:
        from api.security.events import log_security_event
        log_security_event(
            "upload_rejected", None, filename=file.filename, ext=_raw_ext,
            content_type=file.content_type,
        )
        raise HTTPException(
            415,
            f"Unsupported file type '.{_raw_ext or '?'}'. Allowed: "
            f"{', '.join(sorted(ALLOWED_EXTENSIONS))}.",
        )

    file_type = _get_file_type(file.filename or "document.txt", file.content_type or "text/plain")

    # Content-safety: verify the bytes match the claimed type and block
    # zip/XML bombs and disguised binaries before we store or parse the file.
    from api.security.upload_guard import validate_upload_bytes
    validate_upload_bytes(file_bytes, file_type, request)

    # Store a sanitized filename (no path components / unsafe chars) — the
    # download layer re-encodes it, but a clean name is safer everywhere.
    from api.security.sanitize import sanitize_filename
    safe_filename = sanitize_filename(file.filename or "document")
    project_name = name.strip() or safe_filename.rsplit(".", 1)[0]

    # Validate provider name
    valid_providers = {"auto", "openai", "deepl", "azure", "google"}
    if provider_name not in valid_providers:
        provider_name = "auto"

    # Validate language selection against the central language registry
    # (api/languages.py) — adding a new language means adding it there, not here.
    from api.languages import TARGET_LANGUAGES, SOURCE_LANGUAGES
    if target_lang not in TARGET_LANGUAGES:
        raise HTTPException(
            400,
            f"Unsupported target language '{target_lang}'. Supported target languages: {', '.join(TARGET_LANGUAGES)}.",
        )
    if source_lang not in SOURCE_LANGUAGES:
        source_lang = "en"

    # Validate and store layout intelligence config
    _valid_modes = {"original", "saved", "reference"}
    _valid_strengths = {"light", "balanced", "strong"}
    _lc_mode     = layout_mode if layout_mode in _valid_modes else "original"
    _lc_strength = template_strength if template_strength in _valid_strengths else "balanced"
    try:
        _lc_options = json.loads(layout_options or "{}")
        if not isinstance(_lc_options, dict):
            _lc_options = {}
    except Exception:
        _lc_options = {}
    if "image_ocr_enabled" not in _lc_options:
        _lc_options["image_ocr_enabled"] = bool(image_ocr_enabled)
    _lc_options.setdefault("strict_qa", False)
    _lc_options.setdefault("allow_export_with_warnings", True)
    _lc_options.setdefault("auto_repair_enabled", True)
    _lc_options.setdefault("export_best_effort_result", True)

    # Persist the source file on disk instead of as a DB bytea.  Large bytea
    # inserts into PostgreSQL caused connections to be dropped for files in
    # the 15–200 MB range, so we store the path in the DB and keep the bytes
    # on disk.
    source_file_path = save_source_file(project_id, file_bytes, safe_filename)

    project = TranslationProject(
        id=project_id,
        user_id=user["id"],
        name=project_name,
        source_filename=safe_filename,
        source_file_type=file_type,
        source_file_data=None,
        source_file_path=source_file_path,
        source_lang=source_lang,
        target_lang=target_lang,
        style=style,
        keep_english_terms=keep_english_terms,
        transliterate_names=transliterate_names,
        provider_name=provider_name,
        status="ready",
        layout_config={
            "layout_mode":       _lc_mode,
            "style_profile_id":  style_profile_id or "",
            "template_strength": _lc_strength,
            "layout_options":    _lc_options,
        },
    )

    # Read reference template bytes (optional, must be a valid PPTX, max 50 MB)
    if reference_template and reference_template.filename:
        _ref_bytes = await reference_template.read()
        if _ref_bytes and len(_ref_bytes) <= 50 * 1024 * 1024:
            # A reference template is always a .pptx — apply the same content
            # safety check (magic bytes + zip/XML-bomb guard) before storing.
            validate_upload_bytes(_ref_bytes, "pptx", request)
            project.reference_template_data = _ref_bytes
        else:
            log.warning("Reference template ignored (too large or empty): %d bytes", len(_ref_bytes) if _ref_bytes else 0)

    try:
        db.add(project)
        db.commit()
        db.refresh(project)
    except Exception as exc:
        log.exception("Database commit failed for upload user %s (%d bytes): %s", user.get("id"), len(file_bytes), exc)
        delete_source_file(project)
        raise HTTPException(507, f"Storage failure — could not save the uploaded file to the database: {exc}") from exc

    log.info(
        "Upload complete: project %s for user %s, file %s, type %s, %d bytes",
        project.id, user.get("id"), project.source_filename, project.source_file_type, received,
    )

    # NOTE: no document parsing here. Extraction is deferred to the translate
    # step — parsing large decks during upload doubled the wait behind the
    # progress bar, and its estimate was never used by the client (the 428
    # confirmation gate at translate covers cost approval). Upload now costs
    # one file transfer + one DB commit.
    return {
        **_project_summary(project),
        "estimate": None,
        "needs_confirmation": False,
        "message": "Project created. Call the translate endpoint to start translation.",
    }


@router.get("/translation/projects")
def list_projects(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    uid = _user_id(user)
    # Always scope by owner — NULL user_id for anonymous, string id for authenticated
    projects = (
        db.query(TranslationProject)
        .filter(TranslationProject.user_id == uid)
        .order_by(TranslationProject.updated_at.desc())
        .all()
    )
    return [_project_summary(p) for p in projects]


@router.get("/translation/projects/{project_id}")
def get_project(
    project_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    p = _get_owned_project(db, project_id, user)
    return _project_detail(p)


class PatchProjectBody(BaseModel):
    name: Optional[str] = None
    tags: Optional[list[str]] = None
    style: Optional[str] = None
    keep_english_terms: Optional[bool] = None
    transliterate_names: Optional[bool] = None


@router.patch("/translation/projects/{project_id}")
def patch_project(
    project_id: str,
    body: PatchProjectBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    p = _get_owned_project(db, project_id, user)

    if body.name is not None:
        p.name = body.name
    if body.tags is not None:
        p.tags = body.tags
    if body.style is not None:
        p.style = body.style
    if body.keep_english_terms is not None:
        p.keep_english_terms = body.keep_english_terms
    if body.transliterate_names is not None:
        p.transliterate_names = body.transliterate_names

    db.commit()
    return _project_summary(p)


@router.delete("/translation/projects/{project_id}")
def delete_project(
    project_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    p = _get_owned_project(db, project_id, user)
    delete_source_file(p)
    db.delete(p)
    db.commit()
    return {"deleted": project_id}


# ── Duplicate project ─────────────────────────────────────────────────────────

@router.post("/translation/projects/{project_id}/duplicate")
def duplicate_project(
    project_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    p = _get_owned_project(db, project_id, user)
    uid = _user_id(user)
    new_project_id = str(uuid.uuid4())
    new_file_path = copy_source_file(p, new_project_id)
    new_p = TranslationProject(
        id=new_project_id,
        user_id=uid,
        name=f"{p.name} (Copy)",
        source_filename=p.source_filename,
        source_file_type=p.source_file_type,
        source_file_data=None,
        source_file_path=new_file_path,
        source_lang=p.source_lang,
        target_lang=p.target_lang,
        style=p.style,
        keep_english_terms=p.keep_english_terms,
        transliterate_names=p.transliterate_names,
        segments=p.segments[:] if p.segments else [],
        quality_score=p.quality_score,
        quality_issues=p.quality_issues[:] if p.quality_issues else [],
        output_docx=p.output_docx,
        output_pptx=p.output_pptx,
        tags=p.tags[:] if p.tags else [],
        status=p.status,
    )
    db.add(new_p)
    db.commit()
    db.refresh(new_p)
    return _project_summary(new_p)


# ── Provider Settings endpoints ───────────────────────────────────────────────

@router.get("/translation/providers")
def get_providers(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """List all translation providers with their configuration status."""
    uid = _user_id(user)
    return provider_registry.list_providers(db=db, user_id=uid)


class ProviderSettingsBody(BaseModel):
    provider_id: str
    api_key: Optional[str] = None
    is_enabled: Optional[bool] = None
    extra_config: Optional[dict] = None
    max_file_size_mb: Optional[int] = None
    max_pages: Optional[int] = None


class TranslationSettingsBody(BaseModel):
    source_lang: str = "en"
    target_lang: str = "ar"
    style: str = "technical"
    keep_english_terms: bool = False
    transliterate_names: bool = True
    provider_name: str = "auto"
    layout_mode: str = "original"
    style_profile_id: str = ""
    template_strength: str = "balanced"
    layout_options: dict = {}
    # OCR controls are persisted for scanned PDF behavior.
    ocr_enabled: bool = True
    ocr_language: str = "eng"
    ocr_force: bool = False
    # In-image OCR translation for PPTX/DOCX/PDF image content.
    image_ocr_enabled: bool = True
    # Layout QA/export behavior.
    strict_qa: bool = False
    allow_export_with_warnings: bool = True
    auto_repair_enabled: bool = True
    export_best_effort_result: bool = True


def _translation_settings_defaults() -> dict:
    return {
        "source_lang": "en",
        "target_lang": "ar",
        "style": "technical",
        "keep_english_terms": False,
        "transliterate_names": True,
        "provider_name": "auto",
        "layout_mode": "original",
        "style_profile_id": "",
        "template_strength": "balanced",
        "layout_options": {},
        "ocr_enabled": True,
        "ocr_language": "eng",
        "ocr_force": False,
        "image_ocr_enabled": True,
        "strict_qa": False,
        "allow_export_with_warnings": True,
        "auto_repair_enabled": True,
        "export_best_effort_result": True,
    }


def _translation_settings_key(user_id: str | None) -> str:
    return f"translation.settings.{user_id or 'anonymous'}"


def _load_translation_settings(db: Session, user_id: str | None) -> dict:
    from api.db.models import AppSetting

    defaults = _translation_settings_defaults()
    row = db.query(AppSetting).filter(AppSetting.key == _translation_settings_key(user_id)).first()
    if not row or not row.value:
        return defaults

    try:
        parsed = json.loads(row.value)
    except Exception:
        return defaults

    if not isinstance(parsed, dict):
        return defaults

    merged = {**defaults, **parsed}
    if not isinstance(merged.get("layout_options"), dict):
        merged["layout_options"] = {}
    return merged


def _save_translation_settings(db: Session, user_id: str | None, value: dict) -> dict:
    from api.db.models import AppSetting

    defaults = _translation_settings_defaults()
    merged = {**defaults, **value}
    if not isinstance(merged.get("layout_options"), dict):
        merged["layout_options"] = {}

    key = _translation_settings_key(user_id)
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    payload = json.dumps(merged, ensure_ascii=False)

    if row:
        row.value = payload
        row.updated_at = datetime.now(timezone.utc)
    else:
        row = AppSetting(key=key, value=payload)
        db.add(row)
    db.commit()
    return merged


@router.get("/translation/settings")
def get_translation_settings(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Return persisted Translation Studio settings for the current user."""
    uid = _user_id(user)
    return _load_translation_settings(db, uid)


@router.put("/translation/settings")
def update_translation_settings(
    body: TranslationSettingsBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Persist Translation Studio settings for the current user."""
    uid = _user_id(user)
    valid_styles = {"technical", "formal", "bilingual"}
    valid_providers = {"auto", "openai", "deepl", "azure", "google"}
    valid_modes = {"original", "saved", "reference"}
    valid_strengths = {"light", "balanced", "strong"}

    payload = body.model_dump()
    if payload.get("style") not in valid_styles:
        payload["style"] = "technical"
    if payload.get("provider_name") not in valid_providers:
        payload["provider_name"] = "auto"
    if payload.get("layout_mode") not in valid_modes:
        payload["layout_mode"] = "original"
    if payload.get("template_strength") not in valid_strengths:
        payload["template_strength"] = "balanced"
    if payload.get("strict_qa") is None:
        payload["strict_qa"] = False
    if payload.get("allow_export_with_warnings") is None:
        payload["allow_export_with_warnings"] = True
    if payload.get("auto_repair_enabled") is None:
        payload["auto_repair_enabled"] = True
    if payload.get("export_best_effort_result") is None:
        payload["export_best_effort_result"] = True

    return _save_translation_settings(db, uid, payload)


@router.put("/translation/providers")
def update_provider(
    body: ProviderSettingsBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
    _admin: None = Depends(require_admin_session),
):
    """Create or update a provider configuration (admin only — holds API keys)."""
    uid = _user_id(user)
    valid_providers = {"openai", "deepl", "azure", "google"}
    if body.provider_id not in valid_providers:
        raise HTTPException(400, f"Unknown provider '{body.provider_id}'")

    # Find existing config for this user+provider
    cfg = (
        db.query(ProviderConfig)
        .filter(
            ProviderConfig.user_id == uid,
            ProviderConfig.provider_name == body.provider_id,
        )
        .first()
    )
    if cfg is None:
        cfg = ProviderConfig(
            user_id=uid,
            provider_name=body.provider_id,
        )
        db.add(cfg)

    if body.api_key is not None and body.api_key.strip():
        cfg.api_key_enc = encrypt_key(body.api_key.strip())
    if body.is_enabled is not None:
        cfg.is_enabled = body.is_enabled
    if body.extra_config is not None:
        cfg.extra_config = body.extra_config
    if body.max_file_size_mb is not None:
        cfg.max_file_size_mb = body.max_file_size_mb
    if body.max_pages is not None:
        cfg.max_pages = body.max_pages

    db.commit()
    return {"updated": body.provider_id, "is_enabled": cfg.is_enabled}


@router.get("/translation/providers/{provider_id}/health")
async def provider_health(
    provider_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Run a live health check for a provider (latency + connectivity)."""
    uid = _user_id(user)
    valid_providers = {"openai", "deepl", "azure", "google"}
    if provider_id not in valid_providers:
        raise HTTPException(400, f"Unknown provider '{provider_id}'")
    provider = provider_registry.get_provider(provider_id, db=db, user_id=uid)
    result = await provider.health_check()
    return {
        "provider_id": provider_id,
        "is_configured": provider.is_configured,
        **result,
    }


# ── Translation pipeline (SSE) ────────────────────────────────────────────────

@router.post("/translation/projects/{project_id}/translate")
async def translate_project(
    project_id: str,
    request: Request,
    confirmed: bool = False,
    ai_provider: str = "auto",  # "auto" (existing OpenAI/Gemini behavior) | "claude"
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """
    Stream the 6-step translation pipeline as SSE.
    Steps: 1=Extract 2=Memory 3=Prepare 4=Translate 5=Quality 6=Rebuild
    """
    log.info(
        "translate_project start: project_id=%s confirmed=%s user_id=%s",
        project_id,
        confirmed,
        _user_id(user),
    )

    try:
        p = _get_owned_project(db, project_id, user)
        if not has_source_file(p):
            raise HTTPException(400, "No source file uploaded for this project")

        # ── Billing-safety guards (server-side, synchronous — duplicate clicks
        # cannot interleave past them) ─────────────────────────────────────────
        cost_guard.ensure_enabled()

        # Free-tier quota (only when auth is enabled; admins/unlimited bypass).
        # Raises 402 with a sign-in / upgrade prompt when the allowance is spent.
        from api.utils import quota as _quota
        _quota.enforce(db, request, user)

        # Fresh uploads carry no stored segments (extraction is deferred from the
        # upload step to keep uploads fast) — parse once here, reuse in the pipeline.
        _pre_extracted = None
        if not p.segments:
            from api.utils.doc_extractor import extract_document
            from api.languages import ocr_lang_code
            # Language isn't resolved yet when source_lang == "auto" (detection
            # below runs on this very extraction's output) — try every
            # supported OCR language at once rather than assuming English.
            _ocr_lang = (
                ocr_lang_code(p.source_lang)
                if (p.source_lang or "").lower() != "auto"
                else "eng+ara+rus+fra+spa"
            )
            try:
                _pre_extracted = await asyncio.to_thread(
                    extract_document, get_source_bytes(p), p.source_file_type, _ocr_lang
                )
            except Exception as e:
                log.exception(
                    "translate_project parse failed before step 1: project_id=%s user_id=%s",
                    project_id,
                    _user_id(user),
                )
                raise HTTPException(422, f"Could not parse document: {e}")
            # Persist source-only segments so a 428-confirm retry (or a resumed
            # session) can estimate without re-parsing. Small commit — the file
            # bytea column is untouched.
            p.segments = [
                {
                    "id": s["id"],
                    "source": s["source"],
                    "target": "",
                    "seg_type": s.get("seg_type", "paragraph"),
                    "memory_match": False,
                    "flagged": False,
                    "flag_reason": "",
                    "edited": False,
                    "loc": s.get("loc", {}),
                }
                for s in (_pre_extracted or [])
            ]
            db.commit()

        # Resolve "Auto Detect" to a concrete source language before estimating/
        # translating — everything downstream (prompts, RTL layout, filenames,
        # OCR language) keys off a real language code, never the literal "auto".
        if (p.source_lang or "").lower() == "auto":
            from api.languages import SOURCE_LANGUAGES
            sample_text = "\n".join(
                (s.get("source") or "") for s in (p.segments or [])[:20]
            ).strip()
            detected = "en"
            if sample_text:
                try:
                    from langdetect import detect as _detect_lang
                    detected = (_detect_lang(sample_text) or "en").lower().split("-")[0]
                except Exception:
                    log.warning("Language auto-detection failed for project %s; defaulting to English", project_id)
                    detected = "en"
            if detected not in SOURCE_LANGUAGES or detected == "auto":
                detected = "en"
            p.source_lang = detected
            db.commit()
            log.info("Auto-detected source language for project %s: %s", project_id, detected)

        est = cost_guard.estimate_segments(p.segments or [])
        cost_guard.check_job_size(est)
        if cost_guard.needs_confirmation(est) and not confirmed:
            # Client must show "Confirm Translation and Estimated Cost" first,
            # then retry with ?confirmed=true.
            raise HTTPException(428, detail={
                "confirmation_required": True,
                "message": "Large document — confirm the estimated cost to proceed.",
                **est,
            })

        # Kill switch, per-project idempotency (409), concurrency cap (429),
        # per-user hourly rate limit (429), daily/monthly cost ceilings (503).
        # Acquires this project's slot — released in the pipeline's finally.
        cost_guard.check_and_acquire(db, _user_id(user), project_id)
    except HTTPException:
        log.exception(
            "translate_project exited before step 1 (HTTPException): project_id=%s confirmed=%s user_id=%s",
            project_id,
            confirmed,
            _user_id(user),
        )
        raise
    except Exception:
        log.exception(
            "translate_project crashed before step 1: project_id=%s confirmed=%s user_id=%s",
            project_id,
            confirmed,
            _user_id(user),
        )
        raise

    # Capture these values NOW (while the Depends(get_db) session is alive)
    # so the generator can re-open its own session without needing the outer db.
    _project_id = project_id
    _user_id_val = _user_id(user)

    async def _pipeline():
        from openai import AsyncOpenAI
        from api.utils.doc_extractor import extract_document
        from api.utils.translator import translate_segments
        from api.utils.doc_rebuilder import rebuild_document
        from api.languages import ocr_lang_code
        from api.utils.dnt_filter import protect as dnt_protect, restore as dnt_restore, extract_tokens as dnt_extract
        from api.utils.engineering_review import run_engineering_review, compute_consistency_score
        from api.utils.glossary_seed import seed_glossary
        from api.utils.technical_classifier import classify_segments
        from api.db import SessionLocal as _SessionLocal
        from api.db.models import TranslationProject as _TranslationProject

        # 120-second per-request ceiling prevents long request hangs.
        # Initialize lazily and tolerate OpenAI key absence so Gemini-only
        # runtime configurations can still execute the pipeline.
        client = None
        translate_model_name = (os.environ.get("OPENAI_TRANSLATION_MODEL") or "gpt-4o-mini").strip() or "gpt-4o-mini"
        review_model_name = (os.environ.get("OPENAI_REVIEW_MODEL") or "gpt-4o").strip() or "gpt-4o"
        _openai_key = (os.environ.get("OPENAI_API_KEY") or "").strip()
        if ai_provider == "claude":
            from api.services.ai_providers.registry import provider_registry as ai_runtime_registry
            from api.services.ai_providers.claude_openai_compat import ClaudeOpenAICompatClient

            _claude = ai_runtime_registry.get("claude")
            if not _claude or not _claude.is_configured:
                yield _sse({"type": "error", "error": "Claude is not configured (ANTHROPIC_API_KEY missing)."})
                return
            client = ClaudeOpenAICompatClient(api_key=_claude.api_key or os.environ.get("ANTHROPIC_API_KEY", ""))
            translate_model_name = _claude.model or "claude-sonnet-5"
            review_model_name = _claude.model or "claude-sonnet-5"
            log.info("Translation pipeline using Claude (model=%s)", translate_model_name)
        elif _openai_key:
            client = AsyncOpenAI(api_key=_openai_key, timeout=120.0)
        else:
            try:
                from api.services.ai_providers.registry import provider_registry as ai_runtime_registry

                _active_ai = ai_runtime_registry.get_active()
                _active_id = getattr(_active_ai, "provider_id", "") if _active_ai else ""
                _gem_key = (getattr(_active_ai, "api_key", None) or os.environ.get("GEMINI_API_KEY") or "").strip()
                _gem_model = (
                    getattr(_active_ai, "model", None)
                    or os.environ.get("GEMINI_MODEL")
                    or "gemini-3.1-flash-lite"
                )

                if _active_id == "gemini" and _gem_key:
                    client = AsyncOpenAI(
                        api_key=_gem_key,
                        base_url="https://generativelanguage.googleapis.com/v1beta/openai",
                        timeout=120.0,
                    )
                    translate_model_name = _gem_model
                    review_model_name = _gem_model
                    log.info(
                        "Translation pipeline using active Gemini runtime provider (model=%s)",
                        translate_model_name,
                    )
                else:
                    # Final fallback: Claude, if configured. On a deployment where
                    # only ANTHROPIC_API_KEY is set (no OpenAI/Gemini), this lets
                    # the default "auto" engine actually translate instead of
                    # silently having no client. Prefer cheaper OpenAI/Gemini by
                    # setting their keys.
                    from api.services.ai_providers.claude_openai_compat import ClaudeOpenAICompatClient
                    _anthropic_key = (os.environ.get("ANTHROPIC_API_KEY") or "").strip()
                    _claude = ai_runtime_registry.get("claude")
                    if _anthropic_key or (_claude and getattr(_claude, "is_configured", False)):
                        client = ClaudeOpenAICompatClient(
                            api_key=(getattr(_claude, "api_key", None) or _anthropic_key)
                        )
                        _cmodel = (
                            getattr(_claude, "model", None)
                            or os.environ.get("ANTHROPIC_MODEL")
                            or "claude-sonnet-5"
                        )
                        translate_model_name = _cmodel
                        review_model_name = _cmodel
                        log.info("Translation pipeline using Claude fallback (model=%s)", _cmodel)
                    else:
                        log.warning(
                            "No OpenAI/Gemini/Claude provider available for the auto path"
                        )
            except Exception as _client_init_err:
                log.warning("LLM client bootstrap skipped: %s", _client_init_err)

        # ── Billing-safety tracking for this job ────────────────────────────
        _job_t0 = time.time()
        _t_stage: dict[str, float] = {}      # per-stage wall-clock seconds
        _job_usage: dict = {"in": 0, "out": 0, "retries": 0}
        _job_meta: dict = {
            "name": "", "file_type": "", "provider": "",
            "segments_total": 0, "translated": 0, "memory_hits": 0,
            "status": "error", "error": "",
        }

        # ── CRITICAL: open a dedicated session for the pipeline ─────────────
        # Depends(get_db) is torn down when translate_project() returns the
        # StreamingResponse — before this generator runs a single line.
        # Every db operation must go through this self-managed session.
        pipe_db = _SessionLocal()
        _translate_task: asyncio.Task | None = None  # must be set before the try so finally can cancel it safely
        try:
            p = pipe_db.query(_TranslationProject).filter(
                _TranslationProject.id == _project_id,
                _TranslationProject.user_id == _user_id_val,
            ).first()
            if p is None:
                yield _sse({"type": "error", "error": "Project not found"})
                return

            # ── Load layout intelligence config ────────────────────────────────
            _layout_cfg      = p.layout_config or {}
            _layout_mode     = _layout_cfg.get("layout_mode", "original")
            _layout_prof_id  = _layout_cfg.get("style_profile_id", "")
            _tpl_strength    = _layout_cfg.get("template_strength", "balanced")
            _layout_opts     = _layout_cfg.get("layout_options", {})
            if not isinstance(_layout_opts, dict):
                _layout_opts = {}
            _style_profile: dict | None = None
            _strict_qa = bool(_layout_opts.get("strict_qa", False))
            _allow_export_with_warnings = bool(_layout_opts.get("allow_export_with_warnings", True))
            _auto_repair_enabled = bool(_layout_opts.get("auto_repair_enabled", True))
            _export_best_effort_result = bool(_layout_opts.get("export_best_effort_result", True))

            if _layout_mode == "saved" and _layout_prof_id:
                try:
                    from api.db.models import LayoutStyle as _LS
                    _ls = pipe_db.query(_LS).filter(_LS.id == _layout_prof_id).first()
                    if _ls and _ls.properties:
                        _style_profile = _ls.properties
                        log.info(
                            "Translation %s: using saved style '%s' (strength=%s)",
                            _project_id, _ls.name, _tpl_strength,
                        )
                    else:
                        log.warning("Style profile %s not found, using original layout", _layout_prof_id)
                except Exception as _sp_err:
                    log.warning("Could not load style profile: %s", _sp_err)

            elif _layout_mode == "reference" and p.reference_template_data:
                try:
                    from api.services.layout_learner import extract_layout as _extract_layout
                    _style_profile = await asyncio.to_thread(
                        _extract_layout, "reference.pptx", p.reference_template_data
                    )
                    log.info(
                        "Translation %s: using reference template layout (strength=%s)",
                        _project_id, _tpl_strength,
                    )
                except Exception as _ref_err:
                    log.warning("Could not extract reference template layout: %s", _ref_err)

            # Mark project as translating
            p.status = "translating"
            pipe_db.commit()
            _job_meta["name"] = p.name or ""
            _job_meta["file_type"] = p.source_file_type or ""

        except Exception as _init_err:
            log.exception("Pipeline init error")
            cost_guard.release_slot(_project_id)
            try:
                pipe_db.close()
            except Exception:
                pass
            yield _sse({"type": "error", "error": f"Pipeline init failed: {_init_err}"})
            return

        try:
            # ── Step 1: Extract ────────────────────────────────────────────────
            _t_extract_start = time.perf_counter()
            yield _sse({"type": "step", "step": 1, "name": "Extracting document", "status": "running"})
            try:
                if _pre_extracted is not None:
                    # Already parsed in this request's guard block — reuse it
                    raw_segments = _pre_extracted
                else:
                    # CPU-bound parse off the event loop — server stays responsive
                    raw_segments = await asyncio.to_thread(
                        extract_document, get_source_bytes(p), p.source_file_type,
                        ocr_lang_code(p.source_lang),
                    )
                _job_meta["segments_total"] = len(raw_segments or [])
                if not raw_segments:
                    yield _sse({"type": "error", "step": 1, "error": "No translatable text found in document"})
                    p.status = "error"
                    pipe_db.commit()
                    return
            except Exception as e:
                log.exception(
                    "Step 1 extraction failed inside pipeline: project_id=%s user_id=%s",
                    _project_id,
                    _user_id_val,
                )
                yield _sse({"type": "error", "step": 1, "error": f"Extraction failed: {e}"})
                p.status = "error"
                pipe_db.commit()
                return

            # ── Step 1 substep: Seed glossary once per process ──────────────
            # seed_glossary runs 4 DB queries even when already seeded.
            # _GLOSSARY_SEEDED skips the re-check after the first successful seed.
            global _GLOSSARY_SEEDED
            if not _GLOSSARY_SEEDED:
                try:
                    new_terms = seed_glossary(pipe_db)
                    _GLOSSARY_SEEDED = True
                    if new_terms:
                        yield _sse({"type": "substep", "step": 1, "message": f"Seeded {new_terms} domain glossary terms"})
                except Exception as _sg_err:
                    log.warning("Glossary seed failed (non-fatal): %s", _sg_err)

            # ── Step 1 substep: DNT protection ───────────────────────────────
            # Collect all DNT tokens found in source segments
            all_dnt_tokens: list[str] = []
            dnt_maps: list[dict] = []  # one per segment, same order as raw_segments
            for seg in raw_segments:
                source_text = seg.get("source", "")
                tokens = dnt_extract(source_text)
                all_dnt_tokens.extend(tokens)
                protected_text, token_map = dnt_protect(source_text)
                seg["source"] = protected_text
                dnt_maps.append(token_map)

            # Deduplicate
            all_dnt_tokens = list(dict.fromkeys(all_dnt_tokens))

            if all_dnt_tokens:
                yield _sse({
                    "type": "substep", "step": 1,
                    "message": f"Protected {len(all_dnt_tokens)} technical token{'' if len(all_dnt_tokens) == 1 else 's'} (part numbers, codes, units)",
                })

            # Count distinct pages/slides for cost-per-page metric
            _source_pages = len({
                (s.get("loc") or {}).get("slide_idx",
                 (s.get("loc") or {}).get("page_idx",
                  (s.get("loc") or {}).get("page", 0)))
                for s in raw_segments
            }) if raw_segments else 0
            _job_meta["source_pages"] = _source_pages

            _t_stage["extract_s"] = round(time.perf_counter() - _t_extract_start, 2)
            yield _sse({
                "type": "step", "step": 1,
                "name": f"Extracting document ({len(raw_segments)} segments found)",
                "status": "done",
                "data": {
                    "segment_count": len(raw_segments),
                    "elapsed_s": _t_stage["extract_s"],
                },
            })

            # NOTE: Image text detection (diagrams/labels) runs AFTER the main
            # translation pipeline, using gpt-4o-mini, so it cannot exhaust the
            # rate-limit token budget before the primary translation job starts.

            # ── Steps 2–5: Translate ───────────────────────────────────────────
            from api.utils.cost_guard import translation_model as _tm
            _provider_display = {
                "deepl": "DeepL", "azure": "Azure", "google": "Google",
            }.get(p.provider_name or "auto", _tm())

            step_names = [
                None,
                "Extracting",
                "OCR",
                f"Translating via {_provider_display}",
                "Formatting",
                "Quality Check",
                "Rebuilding",
                "Completed",
            ]

            # Real-time progress streaming.
            # translate_segments calls _progress() which puts events in this queue.
            # The generator drains the queue while the task runs, so the client
            # sees each substep immediately instead of waiting for the whole job.
            _progress_queue: asyncio.Queue = asyncio.Queue()

            async def _progress(step: int, total: int, message: str):
                await _progress_queue.put(_sse({
                    "type": "substep",
                    "step": step,
                    "message": message,
                }))

            # ── Incremental save callback ──────────────────────────────────────
            async def _batch_save(current_segments: list) -> None:
                """
                Persist translated segments accumulated so far.
                Sets project status to 'partial' so interrupted jobs are
                visible and resumable in the UI.
                """
                import copy as _copy
                segs_snapshot = [
                    {
                        "id": s["id"],
                        "source": s["source"],
                        "target": s.get("target", ""),
                        "seg_type": s.get("seg_type", "paragraph"),
                        "memory_match": s.get("memory_match", False),
                        "team_match": s.get("team_match", False),
                        "flagged": s.get("flagged", False),
                        "flag_reason": s.get("flag_reason", ""),
                        "edited": False,
                        "loc": s.get("loc", {}),
                    }
                    for s in current_segments
                ]
                p.segments = segs_snapshot
                flag_modified(p, "segments")
                p.status = "partial"
                pipe_db.commit()
                translated_so_far = sum(1 for s in segs_snapshot if s["target"])
                log.info(
                    "Incremental save: %d/%d segments persisted for project %s",
                    translated_so_far, len(segs_snapshot), p.id,
                )

            # ── Step 2: running (memory check) ────────────────────────────────
            yield _sse({"type": "step", "step": 2, "name": step_names[2], "status": "running"})
            _ocr_used = any((s.get("loc") or {}).get("format") == "pdf_ocr" for s in raw_segments)
            if _ocr_used:
                yield _sse({
                    "type": "substep",
                    "step": 2,
                    "message": "OCR complete for scanned PDF pages",
                })
            else:
                yield _sse({
                    "type": "substep",
                    "step": 2,
                    "message": "No OCR needed (selectable text detected)",
                })

            # ── Resolve translation provider ───────────────────────────────────
            # This determines which engine actually translates the document.
            # provider_registry.get_provider() respects the user's selection
            # (or auto-selects the best available configured provider).
            _resolved_provider = None
            _resolved_provider_id = p.provider_name or "auto"
            try:
                _resolved_provider = provider_registry.get_provider(
                    _resolved_provider_id, db=pipe_db, user_id=_user_id_val
                )
                _actual_provider_display = getattr(_resolved_provider, "display_name", _provider_display)
                log.info(
                    "Resolved translation provider: %s → %s (requested: %s)",
                    _resolved_provider_id,
                    _actual_provider_display,
                    p.provider_name,
                )
                # Update SSE label if the resolved provider differs from expected
                if _actual_provider_display != _provider_display:
                    yield _sse({
                        "type": "substep", "step": 3,
                        "message": f"Using {_actual_provider_display} (auto-selected)",
                    })
            except Exception as _prov_err:
                log.warning(
                    "Provider resolution failed (%s) — falling back to OpenAI: %s",
                    _resolved_provider_id, _prov_err,
                )
                _actual_provider_display = "GPT-4o"
                _resolved_provider = None  # translate_segments will use internal path

            # ── Run translation ────────────────────────────────────────────────
            # Launch as a task so we can stream progress events in real time
            # while it runs, instead of blocking until it completes.
            _t_translate_start = time.perf_counter()
            _translate_task = asyncio.create_task(
                translate_segments(
                    segments=raw_segments,
                    source_lang=p.source_lang,
                    target_lang=p.target_lang,
                    style=p.style,
                    keep_english_terms=p.keep_english_terms,
                    transliterate_names=p.transliterate_names,
                    db=pipe_db,
                    user_id=_user_id_val,
                    client=client,
                    progress_callback=_progress,
                    batch_save_callback=_batch_save,
                    provider=_resolved_provider,
                    model_name=translate_model_name,
                    usage=_job_usage,
                    cost_ceiling_usd=cost_guard.config_snapshot()["max_cost_per_job_usd"],
                )
            )

            # Stream progress events while the task runs (0.3 s poll interval)
            while not _translate_task.done():
                try:
                    event = await asyncio.wait_for(_progress_queue.get(), timeout=0.3)
                    yield event
                except asyncio.TimeoutError:
                    pass

            # Drain any events queued in the final moments before task.done() was set
            while not _progress_queue.empty():
                try:
                    yield _progress_queue.get_nowait()
                except asyncio.QueueEmpty:
                    break

            _t_stage["translate_s"] = round(time.perf_counter() - _t_translate_start, 2)

            # Retrieve result — re-raise any exception from inside translate_segments
            _translate_exc = _translate_task.exception()
            if _translate_exc is not None:
                err_msg = str(_translate_exc)
                _job_meta["error"] = err_msg[:512]
                yield _sse({"type": "error", "step": 4, "error": f"Translation failed: {err_msg}"})
                p.status = "error"
                pipe_db.commit()
                return

            translated, quality_score, quality_issues, effective_provider_display, translation_stats = \
                _translate_task.result()
            _job_meta["provider"] = effective_provider_display

            # Count segments that actually got a translation
            actually_translated = sum(1 for s in translated if s.get("target", "").strip())
            _job_meta["translated"] = actually_translated
            _job_meta["memory_hits"] = sum(1 for s in translated if s.get("memory_match"))
            _job_meta["memory_misses"] = len(raw_segments) - _job_meta["memory_hits"]
            _job_meta["retried_segments"] = int(translation_stats.get("retried_segments", 0) or 0)
            _job_meta["failed_segments"] = int(translation_stats.get("failed_segments", 0) or 0)
            _job_meta["failed_details"] = translation_stats.get("failed_details", [])[:50]

            # ── DNT Restoration ─────────────────────────────────────────────────
            # Restore original technical tokens in translated segments.
            # dnt_maps has one entry per raw_segment (in same order).
            dnt_tokens_garbled: list[str] = []
            for i, seg in enumerate(translated):
                if i < len(dnt_maps) and dnt_maps[i]:
                    restored_target, garbled = dnt_restore(seg.get("target", ""), dnt_maps[i])
                    seg["target"] = restored_target
                    dnt_tokens_garbled.extend(garbled)
                    # Also restore source (so segments displayed correctly)
                    restored_source, _ = dnt_restore(seg.get("source", ""), dnt_maps[i])
                    seg["source"] = restored_source

            if dnt_tokens_garbled:
                log.warning("DNT garbled %d tokens: %s", len(dnt_tokens_garbled), dnt_tokens_garbled[:5])

            # ── Emit steps 2–4 summary (with translate timing on step 4) ─────────
            for step_num, status_label in [
                (2, "OCR stage completed"),
                (
                    3,
                    (
                        f"Translated {actually_translated} of {len(translated)} segments via {effective_provider_display}"
                        f" · retries: {_job_meta['retried_segments']}"
                        f" · failed: {_job_meta['failed_segments']}"
                    ),
                ),
                (4, "Formatting translated content"),
            ]:
                extra = {}
                if step_num == 3:
                    extra["elapsed_s"] = _t_stage.get("translate_s", 0)
                yield _sse({
                    "type": "step",
                    "step": step_num,
                    "name": step_names[step_num],
                    "status": "done",
                    "data": {"message": status_label, **extra},
                })

            # ── Step 5: AI Engineering Review (technical segments only) ────────────
            # The technical classifier routes only X-ray/radiation/safety/maintenance
            # segments to GPT-4o review. General segments skip this pass entirely,
            # saving ~60–80% of the engineering review API cost.
            _t_review_start = time.perf_counter()
            yield _sse({"type": "step", "step": 5, "name": step_names[5], "status": "running"})
            engineering_changes: list[dict] = []
            consistency_score = 80
            _expert_reviewed_count = 0
            try:
                _technical_segs, _general_segs = classify_segments(translated)
                _expert_reviewed_count = len(_technical_segs)
                _skipped_review_count = len(_general_segs)

                if _technical_segs:
                    yield _sse({
                        "type": "substep", "step": 5,
                        "message": (
                            f"Expert review: {_expert_reviewed_count} technical segments "
                            f"· {_skipped_review_count} general segments skipped"
                        ),
                    })
                    if client is None:
                        yield _sse({
                            "type": "substep", "step": 5,
                            "message": "Expert review skipped: no LLM client configured for review stage",
                        })
                    else:
                        # Cap the review so it can never hang the whole job (a slow or
                        # failing review provider once stalled a run for ~33 min). Review
                        # is a best-effort enhancement — on timeout the already-translated
                        # segments are kept as-is (engineering_changes stays []).
                        reviewed_technical, engineering_changes = await asyncio.wait_for(
                            run_engineering_review(
                                segments=_technical_segs,
                                source_lang=p.source_lang,
                                target_lang=p.target_lang,
                                client=client,
                                usage=_job_usage,
                                model_name=review_model_name,
                            ),
                            timeout=float(os.environ.get("ENGINEERING_REVIEW_TIMEOUT_S", "120")),
                        )
                        _job_meta["segments_reviewed"] = _expert_reviewed_count
                        # Merge improved translations back — _technical_segs are references
                        # into `translated`, so updating them updates translated in-place.
                        for orig_seg, rev_seg in zip(_technical_segs, reviewed_technical):
                            orig_seg["target"] = rev_seg.get("target", orig_seg.get("target", ""))
                            if rev_seg.get("engineering_reviewed"):
                                orig_seg["engineering_reviewed"] = True
                else:
                    yield _sse({
                        "type": "substep", "step": 5,
                        "message": "No technical segments detected — expert review skipped",
                    })

                consistency_score = compute_consistency_score(translated)
                yield _sse({
                    "type": "substep", "step": 5,
                    "message": (
                        f"Engineering review: {len(engineering_changes)} improvement"
                        f"{'s' if len(engineering_changes) != 1 else ''} applied"
                    ),
                })
                if dnt_tokens_garbled:
                    yield _sse({
                        "type": "substep", "step": 5,
                        "message": f"DNT alert: {len(dnt_tokens_garbled)} token(s) may need review",
                    })
            except Exception as _eng_err:
                log.warning("Engineering review failed (non-fatal): %s", _eng_err)

            _t_stage["review_s"] = round(time.perf_counter() - _t_review_start, 2)

            # ── Compute quality_breakdown with 5 dimensions ─────────────────────
            dnt_total = len(all_dnt_tokens)
            dnt_garbled = len(dnt_tokens_garbled)
            dnt_score = max(0, 100 - int((dnt_garbled / max(dnt_total, 1)) * 100)) if dnt_total > 0 else 100
            formatting_score = min(100, max(0, 100 - len([i for i in quality_issues if i.get("type") == "formatting_error"]) * 10))

            # ── Pipeline cost breakdown ─────────────────────────────────────────
            _total_segs = max(len(translated), 1)
            _memory_hits_count = _job_meta.get("memory_hits", 0)
            _local_rules_count = _job_usage.get("local_rules", 0)
            _provider_translated = max(0, _total_segs - _memory_hits_count - _local_rules_count)
            _pipeline_breakdown = {
                "total_segments": _total_segs,
                "memory_hits": _memory_hits_count,
                "memory_pct": round(_memory_hits_count / _total_segs * 100),
                "local_rules_hits": _local_rules_count,
                "local_rules_pct": round(_local_rules_count / _total_segs * 100),
                "provider_translated": _provider_translated,
                "provider_pct": round(_provider_translated / _total_segs * 100),
                "translated_segments": _job_meta.get("translated", 0),
                "retried_segments": _job_meta.get("retried_segments", 0),
                "failed_segments": _job_meta.get("failed_segments", 0),
                "expert_reviewed": _expert_reviewed_count,
                "expert_reviewed_pct": round(_expert_reviewed_count / _total_segs * 100),
                "skipped_review": _total_segs - _expert_reviewed_count,
                "skipped_review_pct": round((_total_segs - _expert_reviewed_count) / _total_segs * 100),
            }

            quality_breakdown = {
                "translation_quality": quality_score,
                "engineering_quality": min(100, quality_score + min(len(engineering_changes) * 2, 15)),
                "consistency_score": consistency_score,
                "formatting_score": formatting_score,
                "dnt_score": dnt_score,
                "dnt_tokens_found": all_dnt_tokens[:50],
                "dnt_tokens_garbled": dnt_tokens_garbled[:20],
                "engineering_review_changes": len(engineering_changes),
                "provider_used": effective_provider_display,
                "pipeline_breakdown": _pipeline_breakdown,
            }

            yield _sse({
                "type": "substep", "step": 5,
                "message": (
                    f"Pipeline: {_pipeline_breakdown['memory_pct']}% memory · "
                    f"{_pipeline_breakdown['local_rules_pct']}% local rules · "
                    f"{_pipeline_breakdown['provider_pct']}% translated · "
                    f"{_pipeline_breakdown['expert_reviewed_pct']}% expert reviewed"
                ),
            })
            yield _sse({
                "type": "step", "step": 5,
                "name": step_names[5],
                "status": "done",
                "data": {
                    "message": (
                        f"Engineering review complete · Consistency: {consistency_score}% · "
                        f"DNT: {dnt_score}%"
                    ),
                    "pipeline_breakdown": _pipeline_breakdown,
                },
            })

            # ── Step 5.5: Image text detection (runs after main translation) ───
            # Uses gpt-4o-mini (higher rate limit) so it cannot exhaust the
            # token budget before the primary translation job runs.
            if p.source_file_type in ("pdf", "docx", "pptx"):
                yield _sse({
                    "type": "substep",
                    "step": 5,
                    "message": "Scanning embedded diagrams for text labels…",
                })
                try:
                    from openai import AsyncOpenAI as _AsyncOpenAI
                    from api.utils.image_text_extractor import (
                        extract_pdf_images,
                        extract_docx_images,
                        extract_pptx_images,
                        extract_image_labels,
                    )
                    import uuid as _uuid

                    # Dedicated mini client — separate from main translation client so
                    # embedded-diagram OCR uses a cheaper model and doesn't eat into the
                    # primary translation rate-limit budget. Falls back to the already-
                    # resolved `client` (e.g. Gemini's OpenAI-compatible endpoint) when no
                    # OPENAI_API_KEY is configured — previously this always constructed an
                    # unauthenticated OpenAI client for Gemini-only setups, so every
                    # embedded-image OCR call failed silently and no diagram/banner text
                    # was ever translated.
                    _openai_key_present = bool((os.environ.get("OPENAI_API_KEY") or "").strip())
                    if _openai_key_present:
                        mini_client = _AsyncOpenAI(api_key=os.environ.get("OPENAI_API_KEY"))
                        vision_model = "gpt-4o-mini"
                    else:
                        mini_client = client
                        vision_model = translate_model_name

                    if p.source_file_type == "pdf":
                        img_list = extract_pdf_images(get_source_bytes(p))
                    elif p.source_file_type == "docx":
                        img_list = extract_docx_images(get_source_bytes(p))
                    else:
                        img_list = extract_pptx_images(get_source_bytes(p))

                    image_segs_added = 0
                    image_segments: list[dict] = []
                    low_conf_regions = 0
                    class_counts: dict[str, int] = {}
                    image_type_counts: dict[str, int] = {}
                    for img_info in img_list[:20]:  # cap at 20 images to stay within token budget
                        try:
                            if mini_client is None:
                                labels = []
                            else:
                                labels = await extract_image_labels(
                                    img_info["image_bytes"], mini_client, model=vision_model
                                )
                        except Exception as _e:
                            log.warning("Image label extraction error: %s", _e)
                            labels = []
                        for lbl in labels:
                            category = str(lbl.get("category", "other"))
                            image_type = str(lbl.get("image_type", "other"))
                            class_counts[category] = class_counts.get(category, 0) + 1
                            image_type_counts[image_type] = image_type_counts.get(image_type, 0) + 1

                            conf = float(lbl.get("confidence", 0.0) or 0.0)
                            if conf < 0.62:
                                low_conf_regions += 1

                            keep_english = not bool(lbl.get("translate", True))
                            image_segments.append({
                                "id": str(_uuid.uuid4())[:8],
                                "source": lbl["text"],
                                "target": "",
                                "seg_type": "image_text",
                                "memory_match": False,
                                "team_match": False,
                                "flagged": False,
                                "flag_reason": "",
                                "edited": False,
                                "loc": {
                                    "format": p.source_file_type,
                                    "slide_idx": img_info.get("slide_idx"),
                                    "shape_idx": img_info.get("shape_idx"),
                                    "image_index": img_info.get("image_index"),
                                    "page": img_info.get("page"),
                                    "xref": img_info.get("xref"),
                                    "rel_id": img_info.get("rel_id"),
                                    "x_pct": lbl["x_pct"],
                                    "y_pct": lbl["y_pct"],
                                    "bbox": lbl.get("bbox") or {
                                        "x": lbl["x_pct"],
                                        "y": lbl["y_pct"],
                                        "w": 0.10,
                                        "h": 0.05,
                                    },
                                    "text_category": category,
                                    "image_type": image_type,
                                    "ocr_confidence": conf,
                                    "font_size": int(lbl.get("font_size", 14)),
                                    "font_color": str(lbl.get("font_color", "#000000")),
                                    "alignment": str(lbl.get("alignment", "center")),
                                    "keep_english": keep_english,
                                    "passthrough": True,   # skip in coverage gate
                                },
                            })
                            image_segs_added += 1

                    image_segs_translated = 0
                    if image_segments:
                        image_translated, _, _, _, _ = await translate_segments(
                            segments=image_segments,
                            source_lang=p.source_lang,
                            target_lang=p.target_lang,
                            style=p.style,
                            keep_english_terms=p.keep_english_terms,
                            transliterate_names=p.transliterate_names,
                            db=pipe_db,
                            user_id=_user_id_val,
                            client=client,
                            progress_callback=None,
                            batch_save_callback=None,
                            provider=_resolved_provider,
                            model_name=translate_model_name,
                            usage=_job_usage,
                            cost_ceiling_usd=cost_guard.config_snapshot()["max_cost_per_job_usd"],
                        )
                        for seg in image_translated:
                            loc = seg.get("loc", {}) or {}
                            if bool(loc.get("keep_english")):
                                seg["target"] = seg.get("source", "")
                        image_segs_translated = sum(
                            1 for s in image_translated if s.get("target", "").strip()
                        )
                        translated.extend(image_translated)

                    if image_segs_added:
                        top_classes = ", ".join(
                            f"{k}:{v}" for k, v in sorted(class_counts.items(), key=lambda kv: kv[1], reverse=True)[:4]
                        )
                        top_img_types = ", ".join(
                            f"{k}:{v}" for k, v in sorted(image_type_counts.items(), key=lambda kv: kv[1], reverse=True)[:3]
                        )
                        yield _sse({
                            "type": "substep",
                            "step": 5,
                            "message": (
                                f"Found {image_segs_added} text label"
                                f"{'s' if image_segs_added != 1 else ''} in embedded diagrams"
                                f" · translated {image_segs_translated}"
                                f" · low-confidence regions {low_conf_regions}"
                                f" · classes [{top_classes}]"
                                f" · image types [{top_img_types}]"
                            ),
                        })
                    else:
                        yield _sse({
                            "type": "substep",
                            "step": 5,
                            "message": "No text labels detected in embedded images",
                        })
                except Exception as _img_err:
                    log.warning("Image text detection skipped: %s", _img_err)

            # ── Step 6: Rebuild document ───────────────────────────────────────
            _t_rebuild_start = time.perf_counter()
            yield _sse({"type": "step", "step": 6, "name": step_names[6], "status": "running"})

            rebuild_error: str | None = None
            _layout_warnings: list[str] = []  # title overflow/resize notices
            output_docx = output_pptx = output_xlsx = None
            formatting_fidelity = "reconstructed"

            # ── Native Microsoft Office desktop COM backend (preferred) ────────
            # For DOCX/PPTX sources, try the real installed Word/PowerPoint
            # application first — it writes translated text directly into
            # Office's own live paragraph/shape/table objects rather than a
            # python-docx/python-pptx reconstruction, preserving fonts, styles,
            # tables, images, and layout exactly. Falls through to the existing
            # rebuild below on any failure or when Office/COM isn't available;
            # formatting_fidelity only ever says "native_office" when this path
            # actually ran and succeeded.
            if p.source_file_type == "docx":
                from api.utils.word_com_finalizer import (
                    translate_docx_with_word, word_com_available, WordAutomationError,
                )
                if _native_office_enabled() and await asyncio.to_thread(word_com_available):
                    yield _sse({"type": "substep", "step": 6, "message": "Translating in-place via Microsoft Word desktop"})
                    try:
                        output_docx = await asyncio.to_thread(
                            translate_docx_with_word, get_source_bytes(p), translated, p.target_lang
                        )
                        formatting_fidelity = "native_office"
                        yield _sse({"type": "substep", "step": 6, "message": "Microsoft Word native translation complete"})
                    except Exception as _word_err:
                        # Catch ANY failure (COM error OR a bug in the native path)
                        # so the job falls back to the python-docx rebuild.
                        log.warning(
                            "Word native translate failed for project %s, falling back to reconstructed rebuild: %s",
                            _project_id, _word_err, exc_info=True,
                        )
                        output_docx = None
                        formatting_fidelity = "reconstructed"
            elif p.source_file_type == "pptx":
                from api.utils.powerpoint_com_finalizer import (
                    translate_pptx_with_powerpoint, powerpoint_com_available, PowerPointAutomationError,
                )
                if _native_office_enabled() and await asyncio.to_thread(powerpoint_com_available):
                    yield _sse({"type": "substep", "step": 6, "message": "Translating in-place via Microsoft PowerPoint desktop"})
                    try:
                        output_pptx = await asyncio.to_thread(
                            translate_pptx_with_powerpoint, get_source_bytes(p), translated, p.target_lang
                        )
                        formatting_fidelity = "native_office"
                        yield _sse({"type": "substep", "step": 6, "message": "Microsoft PowerPoint native translation complete"})
                    except Exception as _ppt_err:
                        # Catch ANY failure (COM error OR a bug in the native path,
                        # e.g. unexpected segment shape) so the job falls back to
                        # the python-pptx rebuild instead of dying.
                        log.warning(
                            "PowerPoint native translate failed for project %s, falling back to reconstructed rebuild: %s",
                            _project_id, _ppt_err, exc_info=True,
                        )
                        output_pptx = None
                        formatting_fidelity = "reconstructed"

            if formatting_fidelity != "native_office":
                try:
                    output_docx, output_pptx, output_xlsx = rebuild_document(
                        get_source_bytes(p),
                        p.source_file_type,
                        translated,
                        p.target_lang,
                        p.source_lang,
                        layout_warnings=_layout_warnings,
                        style_profile_override=_style_profile,
                        template_strength=_tpl_strength,
                        strict_qa=_strict_qa,
                        allow_export_with_warnings=_allow_export_with_warnings,
                        auto_repair_enabled=_auto_repair_enabled,
                        export_best_effort_result=_export_best_effort_result,
                    )
                except Exception as e:
                    log.error("Document rebuild failed: %s", e)
                    rebuild_error = str(e)
                    output_docx, output_pptx, output_xlsx = None, None, None

            # Surface any Arabic title adjustment notices to the user.
            for _lw in _layout_warnings:
                yield _sse({"type": "substep", "step": 6, "message": _lw})

            # For PPTX / XLSX sources with a successful native rebuild, do NOT
            # fall back to a bilingual DOCX — that would force the wrong format.
            # Only generate a DOCX fallback when no native output exists.
            native_output_ok = bool(output_pptx) or bool(output_xlsx)
            if not output_docx and not native_output_ok:
                try:
                    from api.utils.doc_rebuilder import build_translated_docx
                    output_docx = build_translated_docx(translated, p.source_lang, p.target_lang)
                    formatting_fidelity = "reconstructed"
                    log.info("Used bilingual-fallback DOCX after rebuild failure")
                except Exception as fb_e:
                    log.error("Bilingual fallback also failed: %s", fb_e)
                    output_docx = None

            _t_stage["rebuild_s"] = round(time.perf_counter() - _t_rebuild_start, 2)
            rebuild_ok = bool(output_docx) or bool(output_pptx) or bool(output_xlsx)
            yield _sse({
                "type": "step", "step": 6,
                "name": step_names[6],
                "status": "done" if rebuild_ok else "error",
                "data": {
                    "message": rebuild_error if rebuild_error and not rebuild_ok else "Document rebuilt",
                    "elapsed_s": _t_stage["rebuild_s"],
                    "formatting_fidelity": formatting_fidelity,
                },
            })

            # ── Build segment list for DB ──────────────────────────────────────
            segments_for_db = []
            for s in translated:
                segments_for_db.append({
                    "id": s["id"],
                    "source": s["source"],
                    "target": s.get("target", ""),
                    "seg_type": s.get("seg_type", "paragraph"),
                    "memory_match": s.get("memory_match", False),
                    "team_match": s.get("team_match", False),
                    "flagged": s.get("flagged", False),
                    "flag_reason": s.get("flag_reason", ""),
                    "edited": False,
                    "loc": s.get("loc", {}),
                })

            # ── Step 7: Validate output file ───────────────────────────────────
            _t_validate_start = time.perf_counter()
            yield _sse({"type": "step", "step": 7, "name": step_names[7], "status": "running"})

            has_docx = bool(output_docx and len(output_docx) > 100)
            has_pptx = bool(output_pptx and len(output_pptx) > 100)
            has_xlsx = bool(output_xlsx and len(output_xlsx) > 100)
            has_output = has_docx or has_pptx or has_xlsx

            # Final safety net: only attempt fallback DOCX generation if we
            # have no usable output at all (i.e., not a successful PPTX/XLSX).
            if not has_output:
                try:
                    from api.utils.doc_rebuilder import build_translated_docx
                    output_docx = build_translated_docx(segments_for_db, p.source_lang, p.target_lang)
                    has_docx = bool(output_docx and len(output_docx) > 100)
                    has_output = has_docx
                    if has_docx:
                        p.output_docx = output_docx
                except Exception as final_e:
                    log.error("Final DOCX generation attempt failed: %s", final_e)

            # Determine canonical output format and size for the SSE done event
            if has_pptx:
                output_fmt = "pptx"
                output_bytes_len = len(output_pptx or b'')
            elif has_xlsx:
                output_fmt = "xlsx"
                output_bytes_len = len(output_xlsx or b'')
            else:
                output_fmt = "docx"
                output_bytes_len = len(output_docx or b'')

            # ── Mandatory final layout pass (DOCX only) ────────────────────────
            # Every translated DOCX that did NOT already go through the native
            # translate_docx_with_word path above (which folds this same
            # finalization into its own COM session) is opened in the real
            # layout authority so its RTL/alignment/pagination engine performs
            # the authoritative final formatting pass:
            #   • Windows  → Microsoft Word desktop (COM)
            #   • Linux    → LibreOffice headless
            # This is not optional and has no python-only fallback — if no engine
            # is available or the chosen engine fails, the job fails.
            if output_fmt == "docx" and has_docx and formatting_fidelity != "native_office":
                yield _sse({"type": "substep", "step": 7, "message": "Finalizing DOCX layout (Word / LibreOffice)"})
                from api.utils.document_finalizer import finalize_docx, DocumentFinalizeError
                try:
                    output_docx = await asyncio.to_thread(
                        finalize_docx, output_docx, p.target_lang
                    )
                    output_bytes_len = len(output_docx)
                    has_docx = True
                    yield _sse({"type": "substep", "step": 7, "message": "DOCX finalization complete"})
                except DocumentFinalizeError as _fin_err:
                    log.exception(
                        "DOCX finalization failed for project %s", _project_id
                    )
                    p.status = "error"
                    pipe_db.commit()
                    yield _sse({
                        "type": "error", "step": 7,
                        "error": str(_fin_err),
                    })
                    return

            # Persist final segments + output file before marking complete
            p.segments = segments_for_db
            flag_modified(p, "segments")
            p.quality_score = quality_score
            p.quality_issues = quality_issues
            flag_modified(p, "quality_issues")
            p.quality_breakdown = quality_breakdown
            flag_modified(p, "quality_breakdown")
            p.engineering_review_changes = engineering_changes
            flag_modified(p, "engineering_review_changes")
            p.dnt_tokens = all_dnt_tokens[:100]  # cap for DB storage
            flag_modified(p, "dnt_tokens")
            if output_docx:
                p.output_docx = output_docx
            if output_pptx:
                p.output_pptx = output_pptx
            if output_xlsx:
                p.output_xlsx = output_xlsx
            p.formatting_fidelity = formatting_fidelity

            # ── Layout quality score (PPTX + style profile only) ───────────────
            _layout_quality: dict = {}
            if output_pptx and _style_profile and p.source_file_type == "pptx":
                try:
                    from api.utils.layout_quality_scorer import score_pptx as _score_pptx
                    _layout_quality = await asyncio.to_thread(
                        _score_pptx, output_pptx, _style_profile, p.target_lang
                    )
                    # Persist inside layout_config
                    _updated_cfg = dict(_layout_cfg)
                    _updated_cfg["layout_quality_score"] = _layout_quality
                    p.layout_config = _updated_cfg
                    flag_modified(p, "layout_config")
                    log.info(
                        "Layout quality score for %s: overall=%s font=%s color=%s arabic=%s",
                        _project_id,
                        _layout_quality.get("overall_score"),
                        _layout_quality.get("font_match_pct"),
                        _layout_quality.get("color_match_pct"),
                        _layout_quality.get("arabic_readability_pct"),
                    )
                except Exception as _lq_err:
                    log.warning("Layout quality scoring failed: %s", _lq_err)

            _t_stage["validate_s"] = round(time.perf_counter() - _t_validate_start, 2)
            _t_stage["total_s"] = round(time.time() - _job_t0, 2)
            _warning_count = len(_layout_warnings)
            _slides_requiring_review = sorted({
                int(m.group(1)) for w in _layout_warnings
                for m in [re.search(r"Slide\s+(\d+)", str(w), flags=re.IGNORECASE)]
                if m
            })

            if has_output:
                p.status = "complete"
                _job_meta["status"] = "complete"
                pipe_db.commit()
                _completion_message = (
                    f"Output {output_fmt.upper()} verified ({output_bytes_len:,} bytes)"
                    if _warning_count == 0
                    else (
                        f"Translation completed with layout warnings · "
                        f"{_warning_count} warning(s) · slides requiring review: "
                        f"{', '.join(str(s) for s in _slides_requiring_review[:20]) or 'n/a'}"
                    )
                )
                yield _sse({
                    "type": "step", "step": 7,
                    "name": step_names[7],
                    "status": "done",
                    "data": {
                        "message": _completion_message,
                        "output_fmt": output_fmt,
                        "output_size": output_bytes_len,
                        "elapsed_s": _t_stage["validate_s"],
                        "warnings_count": _warning_count,
                        "slides_requiring_review": _slides_requiring_review,
                        "completion_state": (
                            "completed_with_warnings" if _warning_count > 0 else "completed"
                        ),
                        "download_available": True,
                        "fatal_error": False,
                        "formatting_fidelity": formatting_fidelity,
                        # legacy fields kept for forward-compat with old clients
                        "docx_size": output_bytes_len,
                    },
                })
                # Build cost summary — DeepL is zero-cost (char-based), OpenAI is token-based
                _chars_used = _job_usage.get("chars", 0)
                _cost_summary: dict = {
                    "provider": effective_provider_display,
                    "input_tokens": _job_usage.get("in", 0),
                    "output_tokens": _job_usage.get("out", 0),
                    "est_cost_usd": cost_guard.est_cost_usd(
                        _job_usage.get("in", 0), _job_usage.get("out", 0)
                    ),
                }
                if _chars_used:
                    _cost_summary["deepl_chars"] = _chars_used
                    _cost_summary["deepl_free_remaining"] = max(0, 500_000 - _chars_used)

                # ── Per-stage timing report ────────────────────────────────────
                log.info(
                    "Pipeline timing: extract=%.2fs translate=%.2fs review=%.2fs "
                    "rebuild=%.2fs validate=%.2fs total=%.2fs",
                    _t_stage.get("extract_s", 0),
                    _t_stage.get("translate_s", 0),
                    _t_stage.get("review_s", 0),
                    _t_stage.get("rebuild_s", 0),
                    _t_stage.get("validate_s", 0),
                    _t_stage.get("total_s", 0),
                )

                yield _sse({
                    "type": "done",
                    "project_id": p.id,
                    "segment_count": len(segments_for_db),
                    "translated_segments": _job_meta.get("translated", 0),
                    "retried_segments": _job_meta.get("retried_segments", 0),
                    "failed_segments": _job_meta.get("failed_segments", 0),
                    "quality_score": quality_score,
                    "issue_count": len(quality_issues),
                    "has_output": True,
                    "output_fmt": output_fmt,
                    "output_size": output_bytes_len,
                    "cost_summary": _cost_summary,
                    "timing": _t_stage,
                    "layout_quality": _layout_quality or None,
                    "layout_mode": _layout_mode,
                    "warnings_count": _warning_count,
                    "slides_requiring_review": _slides_requiring_review,
                    "layout_warnings": _layout_warnings,
                    "completion_state": (
                        "completed_with_warnings" if _warning_count > 0 else "completed"
                    ),
                    "download_available": True,
                    "fatal_error": False,
                    "status_message": (
                        "Translation completed with layout warnings"
                        if _warning_count > 0 else "Translation completed"
                    ),
                    # legacy fields
                    "has_docx": has_docx,
                    "has_pptx": has_pptx,
                    "docx_size": output_bytes_len,
                })
            else:
                p.status = "error"
                pipe_db.commit()
                yield _sse({
                    "type": "step", "step": 7,
                    "name": step_names[7],
                    "status": "error",
                    "data": {"message": "Output file could not be generated"},
                })
                yield _sse({
                    "type": "error",
                    "step": 7,
                    "error": (
                        f"Document rebuild failed and fallback also failed. "
                        f"Original error: {rebuild_error or 'unknown'}. "
                        f"The translation segments were saved — you can export manually from the project."
                    ),
                })

        except Exception as e:
            log.exception("Translation pipeline error")
            _job_meta["error"] = str(e)[:512]
            try:
                p.status = "error"
                pipe_db.commit()
            except Exception:
                pass
            yield _sse({"type": "error", "error": str(e)})

        finally:
            # If the client disconnects mid-translation, cancel the background
            # task so it doesn't keep running on a now-closed pipe_db session.
            if _translate_task is not None and not _translate_task.done():
                _translate_task.cancel()
                try:
                    await _translate_task
                except (asyncio.CancelledError, Exception):
                    pass  # expected on cancel

            # Billing safety: always free the concurrency slot and record
            # usage — on success, failure, or client disconnect.
            cost_guard.release_slot(_project_id)
            try:
                cost_guard.record_usage(
                    user_id=_user_id_val,
                    project_id=_project_id,
                    project_name=_job_meta.get("name", ""),
                    file_type=_job_meta.get("file_type", ""),
                    model=cost_guard.translation_model(),
                    provider=_job_meta.get("provider", ""),
                    input_tokens=_job_usage.get("in", 0),
                    output_tokens=_job_usage.get("out", 0),
                    segments_total=_job_meta.get("segments_total", 0),
                    segments_translated=_job_meta.get("translated", 0),
                    memory_hits=_job_meta.get("memory_hits", 0),
                    duration_secs=time.time() - _job_t0,
                    status=_job_meta.get("status", "error"),
                    error=_job_meta.get("error", ""),
                    retries=_job_usage.get("retries", 0),
                    chars_translated=_job_usage.get("chars", 0),
                    translate_in_tokens=_job_usage.get("translate_in", 0),
                    translate_out_tokens=_job_usage.get("translate_out", 0),
                    translate_cached_tokens=_job_usage.get("translate_cached", 0),
                    review_in_tokens=_job_usage.get("review_in", 0),
                    review_out_tokens=_job_usage.get("review_out", 0),
                    review_cached_tokens=_job_usage.get("review_cached", 0),
                    stage_extract_s=_t_stage.get("extract_s", 0),
                    stage_translate_s=_t_stage.get("translate_s", 0),
                    stage_review_s=_t_stage.get("review_s", 0),
                    stage_rebuild_s=_t_stage.get("rebuild_s", 0),
                    stage_validate_s=_t_stage.get("validate_s", 0),
                    api_calls_translate=_job_usage.get("translate_calls", 0),
                    api_calls_review=_job_usage.get("review_calls", 0),
                    segments_reviewed=_job_meta.get("segments_reviewed", 0),
                    memory_misses=_job_meta.get("memory_misses", 0),
                    source_pages=_job_meta.get("source_pages", 0),
                )
            except Exception:
                pass
            try:
                pipe_db.close()
            except Exception:
                pass

    return StreamingResponse(
        _pipeline(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ── Admin: usage & cost dashboard ─────────────────────────────────────────────

@router.get("/translation/admin/usage")
def admin_usage(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Administrator-only API usage / cost dashboard.

    Metadata only — never exposes document contents or API keys.
    Access controlled by the ADMIN_USER_IDS environment variable.
    """
    cost_guard.require_admin(user)

    from sqlalchemy import func as _f
    now = datetime.now(timezone.utc)
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = day_start.replace(day=1)

    def _agg(since):
        q = db.query(
            _f.count(TranslationUsage.id),
            _f.coalesce(_f.sum(TranslationUsage.est_cost_usd), 0.0),
            _f.coalesce(_f.sum(TranslationUsage.input_tokens), 0),
            _f.coalesce(_f.sum(TranslationUsage.output_tokens), 0),
        )
        if since is not None:
            q = q.filter(TranslationUsage.created_at >= since)
        c, cost, tin, tout = q.one()
        return {
            "jobs": int(c or 0),
            "est_cost_usd": round(float(cost or 0), 4),
            "input_tokens": int(tin or 0),
            "output_tokens": int(tout or 0),
        }

    rows = (
        db.query(TranslationUsage)
        .order_by(TranslationUsage.created_at.desc())
        .limit(100)
        .all()
    )
    return {
        "totals": {
            "today": _agg(day_start),
            "month": _agg(month_start),
            "all_time": _agg(None),
        },
        "active_jobs": cost_guard.active_jobs(),
        "config": cost_guard.config_snapshot(),
        "jobs": [
            {
                "id": r.id,
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "user_id": r.user_id,
                "project_id": r.project_id,
                "project_name": r.project_name,
                "file_type": r.file_type,
                "model": r.model,
                "provider": r.provider,
                "input_tokens": r.input_tokens,
                "output_tokens": r.output_tokens,
                "est_cost_usd": r.est_cost_usd,
                "segments_total": r.segments_total,
                "segments_translated": r.segments_translated,
                "memory_hits": r.memory_hits,
                "duration_secs": r.duration_secs,
                "status": r.status,
                "error": r.error,
                "retries": r.retries,
            }
            for r in rows
        ],
    }


@router.post("/translation/track-visit")
def track_visit(request: Request, response: Response, db: Session = Depends(get_db)):
    """Record an anonymous visit (once per browser per UTC day).

    No login required and no personal data stored — a random first-party cookie
    (`ts_vid`) identifies a browser so the admin dashboard can show visitor
    counts. A unique (visitor_id, day) row means each browser counts once daily.
    """
    import uuid as _uuidmod
    from api.config import settings as _settings
    vid = request.cookies.get("ts_vid") or ""
    if not vid or len(vid) != 32 or not all(c in "0123456789abcdef" for c in vid):
        vid = _uuidmod.uuid4().hex
    # Refresh the cookie (1 year). Secure in production; lax so it survives nav.
    response.set_cookie(
        "ts_vid", vid, max_age=31_536_000, httponly=True,
        samesite="strict" if _settings.is_production else "lax",
        secure=_settings.cookie_secure,
    )
    day = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    try:
        from api.db.models import SiteVisit
        db.add(SiteVisit(visitor_id=vid, day=day))
        db.commit()
    except Exception:
        db.rollback()  # already counted this browser today (unique constraint)
    return {"ok": True}


@router.get("/translation/admin/dashboard")
def admin_dashboard(db: Session = Depends(get_db), user: dict = Depends(require_auth)):
    """Simple one-glance admin metrics: visitors, translations, cost vs caps."""
    cost_guard.require_admin(user)
    from sqlalchemy import func as _f
    from api.db.models import SiteVisit

    now = datetime.now(timezone.utc)
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    month_start = day_start.replace(day=1)
    today_str = now.strftime("%Y-%m-%d")

    def _jobs_cost(since):
        q = db.query(
            _f.count(TranslationUsage.id),
            _f.coalesce(_f.sum(TranslationUsage.est_cost_usd), 0.0),
        )
        if since is not None:
            q = q.filter(TranslationUsage.created_at >= since)
        c, cost = q.one()
        return int(c or 0), round(float(cost or 0), 4)

    j_today, c_today = _jobs_cost(day_start)
    j_month, c_month = _jobs_cost(month_start)
    j_all, _ = _jobs_cost(None)

    visitors_today = db.query(_f.count(SiteVisit.id)).filter(SiteVisit.day == today_str).scalar() or 0
    visitors_total = db.query(_f.count(_f.distinct(SiteVisit.visitor_id))).scalar() or 0

    cfg = cost_guard.config_snapshot()
    daily_cap = float(cfg.get("max_daily_api_cost_usd", 0) or 0)
    monthly_cap = float(cfg.get("max_monthly_api_cost_usd", 0) or 0)
    return {
        "visitors": {"today": int(visitors_today), "total": int(visitors_total)},
        "translations": {"today": j_today, "month": j_month, "total": j_all},
        "cost": {
            "today_usd": c_today,
            "month_usd": c_month,
            "daily_cap_usd": daily_cap,
            "monthly_cap_usd": monthly_cap,
            "daily_pct": round(100 * c_today / daily_cap, 1) if daily_cap else 0,
            "monthly_pct": round(100 * c_month / monthly_cap, 1) if monthly_cap else 0,
            "daily_cap_active": not cfg.get("disable_daily_quota", False),
            "monthly_cap_active": not cfg.get("disable_monthly_quota", False),
            "translation_enabled": cfg.get("translation_enabled", True),
        },
        "active_jobs": cost_guard.active_jobs(),
    }


# ── Per-user cost history & chart endpoints ───────────────────────────────────

def _job_row(r) -> dict:
    """Serialize a TranslationUsage row with computed per-stage costs."""
    from api.utils.cost_guard import est_cost_usd

    # Compute stage costs from actual token counts
    _translate_cost = est_cost_usd(
        getattr(r, "translate_in_tokens", 0) or 0,
        getattr(r, "translate_out_tokens", 0) or 0,
        r.model or "gpt-4o-mini",
    )
    _review_cost = est_cost_usd(
        getattr(r, "review_in_tokens", 0) or 0,
        getattr(r, "review_out_tokens", 0) or 0,
        "gpt-4o",
    )
    _total_cost = _translate_cost + _review_cost
    if _total_cost == 0.0:
        _total_cost = float(r.est_cost_usd or 0)

    # Memory savings: memory_hits × avg_cost_per_non_cached_segment
    _mem_hits = getattr(r, "memory_hits", 0) or 0
    _segs_translated = r.segments_translated or 0
    _non_cached = max(_segs_translated - _mem_hits, 0)
    _avg_cost_per_seg = _translate_cost / _non_cached if _non_cached > 0 else 0
    _memory_savings = _mem_hits * _avg_cost_per_seg

    # Cached token savings (50% discount on cached input tokens)
    from api.utils.cost_guard import _PRICES as _P, _price_for as _pf
    _translate_cached = getattr(r, "translate_cached_tokens", 0) or 0
    _review_cached = getattr(r, "review_cached_tokens", 0) or 0
    _pin  = _pf(r.model or "gpt-4o-mini")[0] / 1_000_000   # input price per token
    _rpin = _pf("gpt-4o")[0] / 1_000_000                   # review input price per token
    _cached_savings = _translate_cached * _pin * 0.5 + _review_cached * _rpin * 0.5

    # Review savings: segments skipped from engineering review
    _segs_reviewed = getattr(r, "segments_reviewed", 0) or 0
    _segs_total = r.segments_total or 0
    _review_skipped = max(_segs_total - _segs_reviewed, 0)
    _avg_review_cost_per_seg = _review_cost / _segs_reviewed if _segs_reviewed > 0 else 0
    _review_savings = _review_skipped * _avg_review_cost_per_seg

    return {
        "id": r.id,
        "created_at": r.created_at.isoformat() if r.created_at else None,
        "project_id": r.project_id,
        "project_name": r.project_name,
        "file_type": r.file_type,
        "model": r.model,
        "provider": r.provider,
        "status": r.status,
        "duration_secs": float(r.duration_secs or 0),
        "segments_total": int(r.segments_total or 0),
        "segments_translated": int(r.segments_translated or 0),
        "segments_reviewed": int(getattr(r, "segments_reviewed", 0) or 0),
        "memory_hits": int(r.memory_hits or 0),
        "memory_misses": int(getattr(r, "memory_misses", 0) or 0),
        "source_pages": int(getattr(r, "source_pages", 0) or 0),
        "chars_translated": int(r.chars_translated or 0),
        "retries": int(r.retries or 0),
        # Tokens
        "input_tokens": int(r.input_tokens or 0),
        "output_tokens": int(r.output_tokens or 0),
        "translate_in_tokens": int(getattr(r, "translate_in_tokens", 0) or 0),
        "translate_out_tokens": int(getattr(r, "translate_out_tokens", 0) or 0),
        "translate_cached_tokens": int(getattr(r, "translate_cached_tokens", 0) or 0),
        "review_in_tokens": int(getattr(r, "review_in_tokens", 0) or 0),
        "review_out_tokens": int(getattr(r, "review_out_tokens", 0) or 0),
        "review_cached_tokens": int(getattr(r, "review_cached_tokens", 0) or 0),
        # API calls
        "api_calls_translate": int(getattr(r, "api_calls_translate", 0) or 0),
        "api_calls_review": int(getattr(r, "api_calls_review", 0) or 0),
        # Stage timings
        "stage_extract_s": float(getattr(r, "stage_extract_s", 0) or 0),
        "stage_translate_s": float(getattr(r, "stage_translate_s", 0) or 0),
        "stage_review_s": float(getattr(r, "stage_review_s", 0) or 0),
        "stage_rebuild_s": float(getattr(r, "stage_rebuild_s", 0) or 0),
        "stage_validate_s": float(getattr(r, "stage_validate_s", 0) or 0),
        # Computed costs
        "translate_cost_usd": round(_translate_cost, 6),
        "review_cost_usd": round(_review_cost, 6),
        "total_cost_usd": round(_total_cost, 6),
        "memory_savings_usd": round(_memory_savings, 6),
        "cached_savings_usd": round(_cached_savings, 6),
        "review_savings_usd": round(_review_savings, 6),
    }


@router.get("/translation/cost/history")
def cost_history(
    page: int = 1,
    limit: int = 50,
    sort: str = "created_at",
    order: str = "desc",
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Paginated per-user cost history with full per-stage breakdown.

    Returns only the calling user's own jobs — no admin requirement.
    Sort options: created_at | total_cost_usd | project_name | duration_secs | input_tokens
    """
    uid = _user_id(user)
    q = db.query(TranslationUsage).filter(TranslationUsage.user_id == uid)

    _sort_map = {
        "created_at": TranslationUsage.created_at,
        "total_cost_usd": TranslationUsage.est_cost_usd,
        "project_name": TranslationUsage.project_name,
        "duration_secs": TranslationUsage.duration_secs,
        "input_tokens": TranslationUsage.input_tokens,
    }
    col = _sort_map.get(sort, TranslationUsage.created_at)
    q = q.order_by(col.desc() if order != "asc" else col.asc())

    total = q.count()
    limit = max(1, min(limit, 200))
    offset = (max(1, page) - 1) * limit
    rows = q.offset(offset).limit(limit).all()

    return {
        "jobs": [_job_row(r) for r in rows],
        "total": total,
        "page": page,
        "pages": max(1, -(-total // limit)),  # ceiling division
    }


@router.get("/translation/cost/charts")
def cost_charts(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Aggregated spending charts + savings totals for the calling user."""
    from sqlalchemy import func as _f
    uid = _user_id(user)
    now = datetime.now(timezone.utc)
    day_start   = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start  = day_start - timedelta(days=day_start.weekday())
    month_start = day_start.replace(day=1)

    base_q = db.query(TranslationUsage).filter(TranslationUsage.user_id == uid)

    # ── Summary totals ─────────────────────────────────────────────────────────
    def _sum(since=None, until=None):
        q = db.query(
            _f.count(TranslationUsage.id),
            _f.coalesce(_f.sum(TranslationUsage.est_cost_usd), 0.0),
        ).filter(TranslationUsage.user_id == uid)
        if since:
            q = q.filter(TranslationUsage.created_at >= since)
        if until:
            q = q.filter(TranslationUsage.created_at < until)
        cnt, cost = q.one()
        return int(cnt or 0), round(float(cost or 0), 4)

    today_jobs, today_cost   = _sum(day_start)
    week_jobs,  week_cost    = _sum(week_start)
    month_jobs, month_cost   = _sum(month_start)
    all_jobs,   all_cost     = _sum()

    # ── Daily (last 30 days) ───────────────────────────────────────────────────
    daily = []
    for i in range(29, -1, -1):
        d_start = day_start - timedelta(days=i)
        d_end   = d_start + timedelta(days=1)
        cnt, cost = _sum(d_start, d_end)
        toks = (db.query(_f.coalesce(_f.sum(TranslationUsage.input_tokens), 0))
                .filter(TranslationUsage.user_id == uid,
                        TranslationUsage.created_at >= d_start,
                        TranslationUsage.created_at < d_end)
                .scalar() or 0)
        daily.append({"label": d_start.strftime("%-m/%-d"), "cost": round(float(cost), 4), "jobs": cnt, "tokens": int(toks)})

    # ── Weekly (last 12 weeks) ─────────────────────────────────────────────────
    weekly = []
    for i in range(11, -1, -1):
        w_start = day_start - timedelta(weeks=i, days=day_start.weekday())
        w_end   = w_start + timedelta(weeks=1)
        cnt, cost = _sum(w_start, w_end)
        weekly.append({"label": w_start.strftime("%-m/%-d"), "cost": round(float(cost), 4), "jobs": cnt, "tokens": 0})

    # ── Monthly (last 12 months) ───────────────────────────────────────────────
    monthly = []
    for i in range(11, -1, -1):
        mo = (now.month - i - 1) % 12 + 1
        yr = now.year - ((now.month - i - 1) // 12)
        import calendar as _cal
        m_start = datetime(yr, mo, 1, tzinfo=timezone.utc)
        last_day = _cal.monthrange(yr, mo)[1]
        m_end   = datetime(yr, mo, last_day, 23, 59, 59, tzinfo=timezone.utc)
        cnt, cost = _sum(m_start, m_end)
        monthly.append({"label": m_start.strftime("%b %Y"), "cost": round(float(cost), 4), "jobs": cnt, "tokens": 0})

    # ── Averages ───────────────────────────────────────────────────────────────
    all_rows = base_q.filter(TranslationUsage.est_cost_usd > 0).all()
    avg_cost_per_file = all_cost / all_jobs if all_jobs > 0 else 0.0
    total_pages = sum(getattr(r, "source_pages", 0) or 0 for r in all_rows)
    avg_cost_per_page = all_cost / total_pages if total_pages > 0 else 0.0
    total_words = sum((r.chars_translated or 0) / 5 for r in all_rows)  # chars ÷ 5 ≈ words
    avg_cost_per_1k_words = (all_cost / (total_words / 1000)) if total_words >= 1000 else 0.0

    # ── Savings totals ─────────────────────────────────────────────────────────
    total_memory_savings = 0.0
    total_cached_savings = 0.0
    total_review_savings = 0.0
    for r in all_rows:
        row = _job_row(r)
        total_memory_savings += row["memory_savings_usd"]
        total_cached_savings += row["cached_savings_usd"]
        total_review_savings += row["review_savings_usd"]

    return {
        "daily": daily,
        "weekly": weekly,
        "monthly": monthly,
        "summary": {
            "today_cost": today_cost,  "today_jobs": today_jobs,
            "week_cost":  week_cost,   "week_jobs":  week_jobs,
            "month_cost": month_cost,  "month_jobs": month_jobs,
            "all_cost":   all_cost,    "all_jobs":   all_jobs,
        },
        "avg_cost_per_file":     round(avg_cost_per_file, 6),
        "avg_cost_per_page":     round(avg_cost_per_page, 6),
        "avg_cost_per_1k_words": round(avg_cost_per_1k_words, 6),
        "total_memory_savings":  round(total_memory_savings, 4),
        "total_cached_savings":  round(total_cached_savings, 4),
        "total_review_savings":  round(total_review_savings, 4),
    }


# ── AppSetting helpers ─────────────────────────────────────────────────────────

def _get_setting(db: Session, key: str, default: str = "") -> str:
    from api.db.models import AppSetting
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    return row.value if row else default


def _set_setting(db: Session, key: str, value: str) -> None:
    from api.db.models import AppSetting
    row = db.query(AppSetting).filter(AppSetting.key == key).first()
    if row:
        row.value = value
        row.updated_at = datetime.now(timezone.utc)
    else:
        db.add(AppSetting(key=key, value=value))
    db.commit()


# ── Budget endpoints ───────────────────────────────────────────────────────────

class BudgetBody(BaseModel):
    budget_usd: float


@router.get("/translation/cost/budget")
def get_budget(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Monthly budget status.

    Non-admins receive only their *own* spend this month — no org-wide data.
    Admins receive the full org-wide spend so they can monitor total API costs
    against the configured budget.
    """
    from sqlalchemy import func as _f

    # ── Determine admin status ─────────────────────────────────────────────────
    allowed = {
        a.strip() for a in (os.environ.get("ADMIN_USER_IDS") or "").split(",") if a.strip()
    }
    ident = {str(user.get("id") or ""), str(user.get("email") or "")}
    is_admin_user = bool(allowed and (ident & allowed))

    env_budget = float(os.environ.get("MONTHLY_BUDGET_USD") or "0")
    db_budget = float(_get_setting(db, "monthly_budget_usd", "0") or "0")
    budget_usd = env_budget if env_budget > 0 else db_budget

    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # ── Scope spend: org-wide for admins, personal for everyone else ──────────
    spend_q = db.query(_f.coalesce(_f.sum(TranslationUsage.est_cost_usd), 0.0)).filter(
        TranslationUsage.created_at >= month_start
    )
    if not is_admin_user:
        uid = _user_id(user)
        spend_q = spend_q.filter(TranslationUsage.user_id == uid)

    spent_this_month = float(spend_q.scalar() or 0)

    if budget_usd > 0:
        remaining = max(0.0, budget_usd - spent_this_month)
        pct_consumed = round(min(100.0, (spent_this_month / budget_usd) * 100), 1)
    else:
        remaining = None
        pct_consumed = None

    warning_level = "none"
    if pct_consumed is not None:
        if pct_consumed >= 100:
            warning_level = "critical"
        elif pct_consumed >= 90:
            warning_level = "warn90"
        elif pct_consumed >= 70:
            warning_level = "warn70"

    return {
        "budget_usd": budget_usd,
        "spent_this_month": round(spent_this_month, 4),
        # Label the scope so clients know what they're seeing
        "scope": "org" if is_admin_user else "personal",
        "remaining": round(remaining, 4) if remaining is not None else None,
        "pct_consumed": pct_consumed,
        "warning_level": warning_level,
        "is_admin": is_admin_user,
        "default_markup_pct": float(_get_setting(db, "default_markup_pct", "40") or "40"),
    }


@router.post("/translation/cost/budget")
def set_budget(
    body: BudgetBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
    _admin: None = Depends(require_admin_session),
):
    """Update the monthly OpenAI budget (admin only)."""
    if body.budget_usd < 0:
        raise HTTPException(400, "budget_usd must be >= 0 (0 = unlimited)")
    _set_setting(db, "monthly_budget_usd", str(body.budget_usd))
    return {"ok": True, "budget_usd": body.budget_usd}


# ── Forecast endpoint ──────────────────────────────────────────────────────────

@router.get("/translation/cost/forecast")
def get_forecast(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Cost prediction for the current calendar month — actual completed jobs only."""
    import calendar as _calendar
    from sqlalchemy import func as _f

    uid = _user_id(user)
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    last_day = _calendar.monthrange(now.year, now.month)[1]
    month_end = now.replace(day=last_day, hour=23, minute=59, second=59, microsecond=999999)

    days_elapsed = max(1, (now - month_start).days + 1)
    days_remaining = max(0, (month_end - now).days)
    days_in_month = last_day

    month_cnt, month_cost = db.query(
        _f.count(TranslationUsage.id),
        _f.coalesce(_f.sum(TranslationUsage.est_cost_usd), 0.0),
    ).filter(
        TranslationUsage.user_id == uid,
        TranslationUsage.created_at >= month_start,
    ).one()

    month_cost = float(month_cost or 0)
    month_cnt = int(month_cnt or 0)

    avg_daily_usd = month_cost / days_elapsed
    projected_eom_usd = avg_daily_usd * days_in_month
    avg_cost_per_translation = month_cost / month_cnt if month_cnt > 0 else 0.0

    env_budget = float(os.environ.get("MONTHLY_BUDGET_USD") or "0")
    db_budget = float(_get_setting(db, "monthly_budget_usd", "0") or "0")
    budget_usd = env_budget if env_budget > 0 else db_budget
    remaining_budget_usd = max(0.0, budget_usd - projected_eom_usd) if budget_usd > 0 else None

    return {
        "avg_daily_usd": round(avg_daily_usd, 6),
        "projected_eom_usd": round(projected_eom_usd, 4),
        "remaining_budget_usd": round(remaining_budget_usd, 4) if remaining_budget_usd is not None else None,
        "avg_cost_per_translation": round(avg_cost_per_translation, 6),
        "days_elapsed": days_elapsed,
        "days_remaining": days_remaining,
        "days_in_month": days_in_month,
        "month_cost": round(month_cost, 4),
        "month_jobs": month_cnt,
    }


# ── Customer profitability endpoints ───────────────────────────────────────────

class MarkupBody(BaseModel):
    markup_pct: float


@router.get("/translation/cost/profitability")
def get_profitability(
    markup_pct: float = -1.0,
    page: int = 1,
    limit: int = 50,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Per-job profitability analysis with configurable markup percentage.

    Pass markup_pct=-1 (or omit) to use the admin-configured default.
    """
    uid = _user_id(user)

    if markup_pct < 0:
        stored = _get_setting(db, "default_markup_pct", "40")
        try:
            markup_pct = float(stored)
        except ValueError:
            markup_pct = 40.0

    q = db.query(TranslationUsage).filter(
        TranslationUsage.user_id == uid,
        TranslationUsage.est_cost_usd > 0,
    ).order_by(TranslationUsage.created_at.desc())

    total = q.count()
    limit = max(1, min(limit, 200))
    offset = (max(1, page) - 1) * limit
    rows = q.offset(offset).limit(limit).all()

    jobs = []
    for r in rows:
        actual_cost = float(r.est_cost_usd or 0)
        selling_price = actual_cost * (1 + markup_pct / 100)
        gross_profit = selling_price - actual_cost
        margin_pct = (gross_profit / selling_price * 100) if selling_price > 0 else 0
        jobs.append({
            "job_id": r.id,
            "project_name": r.project_name or "Untitled",
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "file_type": r.file_type or "—",
            "status": r.status or "",
            "segments_total": int(r.segments_total or 0),
            "actual_cost_usd": round(actual_cost, 6),
            "markup_pct": round(markup_pct, 1),
            "selling_price_usd": round(selling_price, 4),
            "gross_profit_usd": round(gross_profit, 4),
            "margin_pct": round(margin_pct, 1),
        })

    total_cost = sum(j["actual_cost_usd"] for j in jobs)
    total_revenue = sum(j["selling_price_usd"] for j in jobs)
    total_profit = sum(j["gross_profit_usd"] for j in jobs)

    return {
        "jobs": jobs,
        "total": total,
        "page": page,
        "pages": max(1, -(-total // limit)),
        "markup_pct": round(markup_pct, 1),
        "summary": {
            "total_cost_usd": round(total_cost, 4),
            "total_revenue_usd": round(total_revenue, 4),
            "total_profit_usd": round(total_profit, 4),
            "avg_margin_pct": round((total_profit / total_revenue * 100) if total_revenue > 0 else 0, 1),
        },
    }


@router.post("/translation/cost/markup")
def set_markup(
    body: MarkupBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
    _admin: None = Depends(require_admin_session),
):
    """Set the default markup percentage (admin only)."""
    if body.markup_pct < 0:
        raise HTTPException(400, "markup_pct must be >= 0")
    _set_setting(db, "default_markup_pct", str(body.markup_pct))
    return {"ok": True, "markup_pct": body.markup_pct}


# ── Live metrics endpoint ──────────────────────────────────────────────────────

@router.get("/translation/cost/live")
def get_live_metrics(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Real-time aggregate metrics — all values derived from actual token columns."""
    from sqlalchemy import func as _f

    uid = _user_id(user)
    now = datetime.now(timezone.utc)
    day_start   = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start  = day_start - timedelta(days=day_start.weekday())
    month_start = day_start.replace(day=1)

    def _cost_since(since) -> float:
        return float(
            db.query(_f.coalesce(_f.sum(TranslationUsage.est_cost_usd), 0.0))
            .filter(TranslationUsage.user_id == uid, TranslationUsage.created_at >= since)
            .scalar() or 0
        )

    today_usd = _cost_since(day_start)
    week_usd  = _cost_since(week_start)
    month_usd = _cost_since(month_start)

    all_rows = db.query(TranslationUsage).filter(
        TranslationUsage.user_id == uid,
        TranslationUsage.est_cost_usd > 0,
    ).all()

    total_cost  = sum(float(r.est_cost_usd or 0) for r in all_rows)
    total_files = len(all_rows)
    total_pages = sum(int(getattr(r, "source_pages", 0) or 0) for r in all_rows)
    total_words = sum(int(r.chars_translated or 0) / 5.0 for r in all_rows)
    total_tokens = sum(
        int(getattr(r, "translate_in_tokens", 0) or 0) +
        int(getattr(r, "translate_out_tokens", 0) or 0) +
        int(getattr(r, "review_in_tokens", 0) or 0) +
        int(getattr(r, "review_out_tokens", 0) or 0)
        for r in all_rows
    )

    return {
        "today_usd":            round(today_usd, 6),
        "week_usd":             round(week_usd, 6),
        "month_usd":            round(month_usd, 6),
        "avg_cost_per_file":    round(total_cost / total_files if total_files else 0, 6),
        "avg_cost_per_page":    round(total_cost / total_pages if total_pages else 0, 6),
        "avg_cost_per_1k_words":round(total_cost / (total_words / 1000) if total_words >= 1000 else 0, 6),
        "avg_cost_per_token":   round(total_cost / total_tokens if total_tokens else 0, 8),
        "total_files":  total_files,
        "total_tokens": total_tokens,
    }


# ── Top-20 most expensive ──────────────────────────────────────────────────────

@router.get("/translation/cost/top20")
def get_top20(
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Top-20 most expensive completed translation jobs by actual API cost."""
    uid = _user_id(user)
    rows = (
        db.query(TranslationUsage)
        .filter(TranslationUsage.user_id == uid, TranslationUsage.est_cost_usd > 0)
        .order_by(TranslationUsage.est_cost_usd.desc())
        .limit(20)
        .all()
    )
    return {
        "jobs": [
            {
                "id": r.id,
                "project_name": r.project_name or "Untitled",
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "est_cost_usd": round(float(r.est_cost_usd or 0), 6),
                "translate_in_tokens":  int(getattr(r, "translate_in_tokens", 0) or 0),
                "translate_out_tokens": int(getattr(r, "translate_out_tokens", 0) or 0),
                "review_in_tokens":     int(getattr(r, "review_in_tokens", 0) or 0),
                "review_out_tokens":    int(getattr(r, "review_out_tokens", 0) or 0),
                "file_type":      r.file_type or "—",
                "model":          r.model or "—",
                "segments_total": int(r.segments_total or 0),
                "source_pages":   int(getattr(r, "source_pages", 0) or 0),
            }
            for r in rows
        ]
    }


# ── Export endpoint ────────────────────────────────────────────────────────────

@router.get("/translation/cost/export")
def export_cost_report(
    format: str = "xlsx",
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Export full job history as Excel (.xlsx), PDF, or CSV."""
    import io
    import csv as _csv_mod
    from fastapi.responses import StreamingResponse

    uid = _user_id(user)
    q = db.query(TranslationUsage).filter(TranslationUsage.user_id == uid)

    if from_date:
        try:
            q = q.filter(TranslationUsage.created_at >= datetime.fromisoformat(from_date))
        except ValueError:
            pass
    if to_date:
        try:
            q = q.filter(TranslationUsage.created_at <= datetime.fromisoformat(to_date))
        except ValueError:
            pass

    rows = q.order_by(TranslationUsage.created_at.desc()).all()
    job_rows = [_job_row(r) for r in rows]

    total_cost    = sum(j["total_cost_usd"] for j in job_rows)
    total_tokens  = sum(
        j["translate_in_tokens"] + j["translate_out_tokens"] +
        j["review_in_tokens"] + j["review_out_tokens"] for j in job_rows
    )
    total_savings = sum(
        j["memory_savings_usd"] + j["cached_savings_usd"] + j["review_savings_usd"]
        for j in job_rows
    )

    col_headers = [
        "Date", "Project Name", "File Type", "Status", "Model",
        "Segments", "Pages",
        "Translate In Tokens", "Translate Out Tokens",
        "Review In Tokens", "Review Out Tokens",
        "Translate Cost (USD)", "Review Cost (USD)", "Total Cost (USD)",
        "Memory Savings (USD)", "Cache Savings (USD)", "Review Savings (USD)",
        "Duration (s)",
    ]

    def _vals(j):
        return [
            (j["created_at"] or "")[:19].replace("T", " "),
            j["project_name"] or "",
            j["file_type"] or "",
            j["status"] or "",
            j["model"] or "",
            j["segments_total"],
            j["source_pages"],
            j["translate_in_tokens"],
            j["translate_out_tokens"],
            j["review_in_tokens"],
            j["review_out_tokens"],
            j["translate_cost_usd"],
            j["review_cost_usd"],
            j["total_cost_usd"],
            j["memory_savings_usd"],
            j["cached_savings_usd"],
            j["review_savings_usd"],
            round(j["duration_secs"], 1),
        ]

    # ── CSV ───────────────────────────────────────────────────────────────────
    if format == "csv":
        buf = io.StringIO()
        w = _csv_mod.writer(buf)
        w.writerow(["TRANSLATION COST REPORT"])
        w.writerow(["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
        w.writerow(["Total Jobs", len(job_rows)])
        w.writerow(["Total Cost (USD)", f"{total_cost:.6f}"])
        w.writerow(["Total Tokens", total_tokens])
        w.writerow(["Total Savings (USD)", f"{total_savings:.6f}"])
        w.writerow([])
        w.writerow(col_headers)
        for j in job_rows:
            w.writerow(_vals(j))
        content = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=translation-cost-report.csv"},
        )

    # ── XLSX ──────────────────────────────────────────────────────────────────
    elif format == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cost Report"

        header_font  = Font(bold=True, color="FFFFFF", size=11)
        header_fill  = PatternFill("solid", fgColor="1E3A5F")
        summary_font = Font(bold=True)
        alt_fill     = PatternFill("solid", fgColor="F8FAFC")
        money_fmt    = '#,##0.000000'
        center_al    = Alignment(horizontal="center", vertical="center")

        summary_block = [
            ("TRANSLATION COST REPORT", ""),
            ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Jobs", len(job_rows)),
            ("Total Cost (USD)", total_cost),
            ("Total Tokens", total_tokens),
            ("Total Savings (USD)", total_savings),
        ]
        for ri, (lbl, val) in enumerate(summary_block, 1):
            ws.cell(ri, 1, lbl).font = summary_font
            c = ws.cell(ri, 2, val)
            if ri in (4, 6):
                c.number_format = money_fmt

        hrow = len(summary_block) + 2
        for ci, h in enumerate(col_headers, 1):
            cell = ws.cell(hrow, ci, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_al

        for ri, j in enumerate(job_rows, hrow + 1):
            vals = _vals(j)
            is_alt = (ri - hrow) % 2 == 0
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(ri, ci, v)
                if is_alt:
                    cell.fill = alt_fill
                if 12 <= ci <= 17:        # cost/savings columns
                    cell.number_format = money_fmt

        for col in ws.columns:
            max_w = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 4, 42)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=translation-cost-report.xlsx"},
        )

    # ── PDF ───────────────────────────────────────────────────────────────────
    elif format == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            topMargin=1.5*cm, bottomMargin=1.5*cm,
            leftMargin=1.5*cm, rightMargin=1.5*cm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "rptTitle", parent=styles["Heading1"], fontSize=16, spaceAfter=4,
            textColor=colors.HexColor("#1E3A5F"),
        )
        sub_style = ParagraphStyle(
            "rptSub", parent=styles["Normal"], fontSize=9, spaceAfter=0,
            textColor=colors.HexColor("#475569"),
        )

        story: list = [
            Paragraph("Translation Cost Report", title_style),
            Paragraph(
                f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} "
                f" | Jobs: {len(job_rows)}"
                f" | Total Cost: ${total_cost:.4f}"
                f" | Total Savings: ${total_savings:.4f}",
                sub_style,
            ),
            Spacer(1, 0.5*cm),
        ]

        short_hdrs = ["Date", "Project", "Type", "Model", "Segs", "Pgs",
                      "Translate $", "Review $", "Total $", "Savings $", "Dur(s)"]
        col_w      = [2.8*cm, 5*cm, 1.4*cm, 2.4*cm, 1.2*cm, 1.1*cm,
                      2.2*cm, 2.2*cm, 2.2*cm, 2.2*cm, 1.4*cm]

        tdata = [short_hdrs]
        for j in job_rows:
            sav = j["memory_savings_usd"] + j["cached_savings_usd"] + j["review_savings_usd"]
            tdata.append([
                (j["created_at"] or "")[:10],
                (j["project_name"] or "")[:36],
                (j["file_type"] or "")[:6],
                (j["model"] or "")[:14],
                str(j["segments_total"]),
                str(j["source_pages"]),
                f"${j['translate_cost_usd']:.4f}",
                f"${j['review_cost_usd']:.4f}",
                f"${j['total_cost_usd']:.4f}",
                f"${sav:.4f}",
                str(round(j["duration_secs"], 1)),
            ])

        t = Table(tdata, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, -1), 7),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
            ("ALIGN",         (4, 1), (-1, -1), "RIGHT"),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=translation-cost-report.pdf"},
        )

    else:
        raise HTTPException(400, f"Unknown format '{format}'. Use xlsx, pdf, or csv.")


# ── Segment editing ───────────────────────────────────────────────────────────

class SegmentEditBody(BaseModel):
    target: str


@router.patch("/translation/projects/{project_id}/segments/{segment_id}")
def edit_segment(
    project_id: str,
    segment_id: str,
    body: SegmentEditBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Edit the translated text of a single segment and save to translation memory."""
    import copy as _copy
    p = _get_owned_project(db, project_id, user)

    # Deep-copy so SQLAlchemy detects the mutation as a new object
    segments = _copy.deepcopy(p.segments or [])
    updated = False
    for seg in segments:
        if seg["id"] == segment_id:
            # Save to translation memory
            if seg.get("source") and body.target:
                from api.utils.translator import _save_to_memory
                _save_to_memory(db, user["id"], seg["source"], body.target, p.source_lang, p.target_lang)

            seg["target"] = body.target
            seg["edited"] = True
            seg["flagged"] = False
            seg["flag_reason"] = ""
            updated = True
            break

    if not updated:
        raise HTTPException(404, "Segment not found")

    p.segments = segments
    flag_modified(p, "segments")
    db.commit()
    return {"updated": segment_id, "target": body.target}


# ── Version management ────────────────────────────────────────────────────────

class SaveVersionBody(BaseModel):
    name: Optional[str] = None


@router.post("/translation/projects/{project_id}/versions")
def save_version(
    project_id: str,
    body: SaveVersionBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Snapshot the current translation state as a named version."""
    import copy as _copy
    p = _get_owned_project(db, project_id, user)

    # Deep-copy to create a new object so SQLAlchemy detects the mutation
    versions = _copy.deepcopy(p.versions or [])
    new_vn = max((v["version_num"] for v in versions), default=0) + 1
    versions.append({
        "version_num": new_vn,
        "name": body.name or f"Version {new_vn}",
        "created_at": _now_iso(),
        "quality_score": p.quality_score,
        "segment_count": len(p.segments or []),
        "segments": _copy.deepcopy(p.segments or []),
    })
    p.versions = versions
    flag_modified(p, "versions")
    p.version_num = new_vn
    db.commit()
    return {"version_num": new_vn, "name": versions[-1]["name"]}


@router.post("/translation/projects/{project_id}/versions/{version_num}/restore")
def restore_version(
    project_id: str,
    version_num: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Restore a previous version of the translation."""
    import copy as _copy
    p = _get_owned_project(db, project_id, user)

    versions = p.versions or []
    version = next((v for v in versions if v["version_num"] == version_num), None)
    if not version:
        raise HTTPException(404, f"Version {version_num} not found")

    # Deep-copy the stored segments so the restored object is a distinct instance
    p.segments = _copy.deepcopy(version.get("segments", []))
    flag_modified(p, "segments")
    p.quality_score = version.get("quality_score")
    db.commit()
    return {"restored": version_num, "segment_count": len(p.segments)}


# ── Export ────────────────────────────────────────────────────────────────────

def _coverage_report(segments: list[dict]) -> dict:
    """
    Analyse translation coverage for a segment list.

    Returns a dict with:
      total        — total translatable segments (excludes passthroughs)
      translated   — segments that have a non-empty target differing from source
      passthroughs — segments whose loc marks them as passthrough (kept as-is)
      pct          — translated / max(total,1) as a 0-100 float
      untranslated — list of {id, source} for segments with no target (first 20)
    """
    translatable = []
    passthroughs = 0
    for seg in segments:
        if seg.get("loc", {}).get("passthrough"):
            passthroughs += 1
            continue
        translatable.append(seg)

    translated = []
    untranslated = []
    for seg in translatable:
        src = seg.get("source", "").strip()
        tgt = seg.get("target", "").strip()
        # Non-empty target = translated. target==source is valid when the
        # translation decision is to keep a term in the source language
        # (abbreviations, part numbers, model codes, proper nouns).
        # Empty target = pipeline produced nothing for this segment.
        if tgt:
            translated.append(seg)
        else:
            untranslated.append({"id": seg.get("id"), "source": src[:120]})

    total = len(translatable)
    pct = round(len(translated) / max(total, 1) * 100, 1)
    return {
        "total": total,
        "translated": len(translated),
        "passthroughs": passthroughs,
        "pct": pct,
        "untranslated": untranslated[:20],
    }


@router.get("/translation/projects/{project_id}/export/{fmt}")
def export_project(
    project_id: str,
    fmt: str,
    rebuild: bool = False,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Download the translated document in the requested format.

    Pass ?rebuild=true to discard the cached PPTX/DOCX/XLSX and produce a
    fresh file from the stored segments using the current pipeline.  The
    original translation is preserved — only the document assembly step runs
    again.  Useful after a pipeline fix to pick up corrected layout behaviour
    without paying for translation again.
    """
    p = _get_owned_project(db, project_id, user)
    if not p.segments:
        raise HTTPException(400, "No translated content available")

    # Force-rebuild: clear the cached native output so the code below always
    # rebuilds from segments.  We do NOT clear output_docx fallbacks here
    # because those are generated differently; only clear the format requested.
    if rebuild:
        fmt_lower = fmt.lower()
        if fmt_lower == "pptx" and p.output_pptx:
            log.info("?rebuild=true: clearing cached PPTX for project %s", project_id)
            p.output_pptx = None
            db.commit()
        elif fmt_lower == "xlsx" and getattr(p, "output_xlsx", None):
            log.info("?rebuild=true: clearing cached XLSX for project %s", project_id)
            p.output_xlsx = None
            db.commit()
        elif fmt_lower == "docx" and getattr(p, "output_docx", None):
            log.info("?rebuild=true: clearing cached DOCX for project %s", project_id)
            p.output_docx = None
            db.commit()

    fmt = fmt.lower()

    if fmt == "arabic":
        # Alias for language-specific export while preserving native document type.
        # A PDF source downloads as PDF (not DOCX) — see the "pdf" branch, which
        # converts the translated DOCX to PDF via LibreOffice.
        fmt = p.source_file_type if p.source_file_type in ("pptx", "xlsx", "docx", "pdf") else "docx"

    if fmt == "original":
        if not has_source_file(p):
            raise HTTPException(400, "No original file available")
        original_bytes = get_source_bytes(p)
        source_ext = (p.source_filename.rsplit(".", 1)[-1].lower() if "." in p.source_filename else p.source_file_type)
        return Response(
            content=original_bytes,
            media_type=mime_for_ext(source_ext),
            headers={
                "Content-Disposition": content_disposition(p.source_filename or f"original.{source_ext}"),
                "Content-Length": str(len(original_bytes or b"")),
                "Cache-Control": "private, no-store",
            },
        )
    safe_name = p.name.replace(" ", "_").replace("/", "_")[:60]  # kept for internal use only

    from api.utils.filename_helper import build_translated_filename_from_code
    dl_filename = build_translated_filename_from_code(p.source_filename, p.target_lang)

    # ── Coverage gate (DOCX / PPTX in-place rebuilds) ─────────────────────────
    # Bilingual fallback formats (txt, html, pdf, bilingual docx) always render
    # source text for untranslated segments, so the gate only blocks structured
    # rebuild formats where missing Arabic would leave English holes in the file.
    if fmt in ("docx", "pptx"):
        report = _coverage_report(p.segments)
        COVERAGE_THRESHOLD = 95.0
        if report["pct"] < COVERAGE_THRESHOLD:
            untranslated_list = "\n".join(
                f'  [{i+1}] {u["source"]}'
                for i, u in enumerate(report["untranslated"])
            )
            extra = ""
            if len(report["untranslated"]) < (report["total"] - report["translated"]):
                extra = f"\n  … and {report['total'] - report['translated'] - len(report['untranslated'])} more"
            raise HTTPException(
                status_code=422,
                detail={
                    "error": "translation_coverage_too_low",
                    "message": (
                        f"Translation coverage is {report['pct']}% "
                        f"({report['translated']}/{report['total']} segments). "
                        f"A minimum of {int(COVERAGE_THRESHOLD)}% is required before export."
                    ),
                    "coverage_pct": report["pct"],
                    "translated": report["translated"],
                    "total": report["total"],
                    "untranslated_count": report["total"] - report["translated"],
                    "sample_untranslated": report["untranslated"],
                },
            )

    if fmt == "english":
        # Build an English-only export by reusing source text in translated slots.
        english_segments = []
        for seg in (p.segments or []):
            cloned = dict(seg)
            cloned["target"] = seg.get("source", "")
            english_segments.append(cloned)

        try:
            if p.source_file_type == "pptx" and has_source_file(p):
                from api.utils.doc_rebuilder import rebuild_pptx

                content = rebuild_pptx(get_source_bytes(p), english_segments, "en")
                dl_name = build_translated_filename_from_code(p.source_filename, "en")
                if not dl_name.lower().endswith(".pptx"):
                    dl_name = (dl_name.rsplit(".", 1)[0] + ".pptx") if "." in dl_name else (dl_name + ".pptx")
                return Response(
                    content=content,
                    media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    headers={
                        "Content-Disposition": content_disposition(dl_name),
                        "Content-Length": str(len(content)),
                        "Cache-Control": "private, no-store",
                    },
                )

            if p.source_file_type == "xlsx" and has_source_file(p):
                from api.utils.doc_rebuilder import rebuild_xlsx

                content = rebuild_xlsx(get_source_bytes(p), english_segments, "en")
                dl_name = build_translated_filename_from_code(p.source_filename, "en")
                if not dl_name.lower().endswith(".xlsx"):
                    dl_name = (dl_name.rsplit(".", 1)[0] + ".xlsx") if "." in dl_name else (dl_name + ".xlsx")
                return Response(
                    content=content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": content_disposition(dl_name),
                        "Content-Length": str(len(content)),
                        "Cache-Control": "private, no-store",
                    },
                )

            if p.source_file_type == "docx" and has_source_file(p):
                from api.utils.doc_rebuilder import rebuild_docx

                content = rebuild_docx(get_source_bytes(p), english_segments, "en")
            else:
                from api.utils.doc_rebuilder import build_translated_docx

                content = build_translated_docx(english_segments, p.source_lang, "en")
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(500, f"Could not build English export: {exc}")

        content = _word_finalize_or_503(content, "en")

        dl_name = build_translated_filename_from_code(p.source_filename, "en")
        if not dl_name.lower().endswith(".docx"):
            dl_name = (dl_name.rsplit(".", 1)[0] + ".docx") if "." in dl_name else (dl_name + ".docx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={
                "Content-Disposition": content_disposition(dl_name),
                "Content-Length": str(len(content)),
                "Cache-Control": "private, no-store",
            },
        )

    if fmt == "package":
        if not has_source_file(p):
            raise HTTPException(400, "No source file available for package export")

        from api.utils.doc_rebuilder import rebuild_docx, rebuild_pptx, rebuild_xlsx, build_translated_docx

        arabic_name = build_translated_filename_from_code(p.source_filename, p.target_lang)
        english_name = build_translated_filename_from_code(p.source_filename, "en")
        original_name = p.source_filename or "original"

        english_segments = []
        for seg in (p.segments or []):
            cloned = dict(seg)
            cloned["target"] = seg.get("source", "")
            english_segments.append(cloned)

        src_bytes = get_source_bytes(p)
        arabic_bytes = None
        english_bytes = None

        if p.source_file_type == "pptx":
            arabic_bytes = p.output_pptx or rebuild_pptx(src_bytes, p.segments, p.target_lang)
            english_bytes = rebuild_pptx(src_bytes, english_segments, "en")
            if not arabic_name.lower().endswith(".pptx"):
                arabic_name = (arabic_name.rsplit(".", 1)[0] + ".pptx") if "." in arabic_name else (arabic_name + ".pptx")
            if not english_name.lower().endswith(".pptx"):
                english_name = (english_name.rsplit(".", 1)[0] + ".pptx") if "." in english_name else (english_name + ".pptx")
        elif p.source_file_type == "xlsx":
            arabic_bytes = p.output_xlsx or rebuild_xlsx(src_bytes, p.segments, p.target_lang)
            english_bytes = rebuild_xlsx(src_bytes, english_segments, "en")
            if not arabic_name.lower().endswith(".xlsx"):
                arabic_name = (arabic_name.rsplit(".", 1)[0] + ".xlsx") if "." in arabic_name else (arabic_name + ".xlsx")
            if not english_name.lower().endswith(".xlsx"):
                english_name = (english_name.rsplit(".", 1)[0] + ".xlsx") if "." in english_name else (english_name + ".xlsx")
        else:
            if p.source_file_type == "docx":
                if p.output_docx:
                    arabic_bytes = p.output_docx  # already Word-finalized by the pipeline
                else:
                    arabic_bytes = _word_finalize_or_503(
                        rebuild_docx(src_bytes, p.segments, p.target_lang), p.target_lang
                    )
                english_bytes = _word_finalize_or_503(
                    rebuild_docx(src_bytes, english_segments, "en"), "en"
                )
            else:
                if p.output_docx:
                    arabic_bytes = p.output_docx  # already Word-finalized by the pipeline
                else:
                    arabic_bytes = _word_finalize_or_503(
                        build_translated_docx(p.segments, p.source_lang, p.target_lang), p.target_lang
                    )
                english_bytes = _word_finalize_or_503(
                    build_translated_docx(english_segments, p.source_lang, "en"), "en"
                )
            if not arabic_name.lower().endswith(".docx"):
                arabic_name = (arabic_name.rsplit(".", 1)[0] + ".docx") if "." in arabic_name else (arabic_name + ".docx")
            if not english_name.lower().endswith(".docx"):
                english_name = (english_name.rsplit(".", 1)[0] + ".docx") if "." in english_name else (english_name + ".docx")

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"original/{original_name}", src_bytes)
            zf.writestr(f"arabic/{arabic_name}", arabic_bytes or b"")
            zf.writestr(f"english/{english_name}", english_bytes or b"")

        package_name = build_translated_filename_from_code(p.source_filename, p.target_lang)
        package_base = package_name.rsplit(".", 1)[0] if "." in package_name else package_name
        package_file = f"{package_base} (Arabic+English).zip"
        zip_bytes = zip_buf.getvalue()
        return Response(
            content=zip_bytes,
            media_type="application/zip",
            headers={
                "Content-Disposition": content_disposition(package_file),
                "Content-Length": str(len(zip_bytes)),
                "Cache-Control": "private, no-store",
            },
        )

    if fmt == "docx":
        if p.output_docx:
            # Stored in-place rebuild from the translation pipeline — already
            # passed through the mandatory Word finalization pass.
            content = p.output_docx
        elif p.source_file_type == "docx" and has_source_file(p):
            # Pipeline rebuild failed or was skipped — retry on-the-fly.
            # Prefer the native Word COM backend (same as the pipeline); fall
            # back to the python-docx rebuild + mandatory finalize on failure.
            from api.utils.word_com_finalizer import (
                translate_docx_with_word, word_com_available, WordAutomationError,
            )
            content = None
            _export_fidelity = "reconstructed"
            if _native_office_enabled() and word_com_available():
                try:
                    content = translate_docx_with_word(get_source_bytes(p), p.segments, p.target_lang)
                    _export_fidelity = "native_office"
                except Exception as _word_err:
                    log.warning("On-the-fly Word native translate failed (%s), falling back to reconstructed rebuild", _word_err, exc_info=True)
                    content = None
            if content is None:
                try:
                    from api.utils.doc_rebuilder import rebuild_docx
                    content = rebuild_docx(get_source_bytes(p), p.segments, p.target_lang)
                except Exception as _e:
                    log.warning("On-the-fly rebuild_docx failed (%s), falling back to bilingual", _e)
                    from api.utils.doc_rebuilder import build_translated_docx
                    content = build_translated_docx(p.segments, p.source_lang, p.target_lang)
                content = _word_finalize_or_503(content, p.target_lang)
            # Cache the finalized bytes so future downloads are instant
            p.output_docx = content
            p.formatting_fidelity = _export_fidelity
            db.commit()
        else:
            # PDF / TXT / HTML source — use bilingual two-column output
            from api.utils.doc_rebuilder import build_translated_docx
            content = build_translated_docx(p.segments, p.source_lang, p.target_lang)
            content = _word_finalize_or_503(content, p.target_lang)
        dl_name_docx = dl_filename if dl_filename.lower().endswith(".docx") else (dl_filename.rsplit(".", 1)[0] + ".docx" if "." in dl_filename else dl_filename + ".docx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={
                "Content-Disposition": content_disposition(dl_name_docx),
                "Content-Length": str(len(content)),
                "Cache-Control": "private, no-store",
            },
        )

    elif fmt == "pptx":
        from api.utils.doc_rebuilder import rebuild_pptx, validate_pptx_bytes

        _layout_cfg = p.layout_config or {}
        _layout_mode = _layout_cfg.get("layout_mode", "original")
        _layout_prof_id = _layout_cfg.get("style_profile_id", "")
        _tpl_strength = _layout_cfg.get("template_strength", "balanced")
        _layout_opts = _layout_cfg.get("layout_options", {})
        if not isinstance(_layout_opts, dict):
            _layout_opts = {}

        _strict_qa = bool(_layout_opts.get("strict_qa", False))
        _allow_export_with_warnings = bool(_layout_opts.get("allow_export_with_warnings", True))
        _auto_repair_enabled = bool(_layout_opts.get("auto_repair_enabled", True))
        _export_best_effort_result = bool(_layout_opts.get("export_best_effort_result", True))

        _style_profile: dict | None = None
        if _layout_mode == "saved" and _layout_prof_id:
            try:
                from api.db.models import LayoutStyle as _LS

                _ls = db.query(_LS).filter(_LS.id == _layout_prof_id).first()
                if _ls and _ls.properties:
                    _style_profile = _ls.properties
            except Exception as _style_err:
                log.warning("Could not load style profile for export rebuild: %s", _style_err)
        elif _layout_mode == "reference" and p.reference_template_data:
            try:
                from api.services.layout_learner import extract_layout as _extract_layout

                _style_profile = _extract_layout("reference.pptx", p.reference_template_data)
            except Exception as _ref_err:
                log.warning("Could not extract reference template for export rebuild: %s", _ref_err)

        log.info(
            "PPTX export requested: project=%s user=%s cached=%s source_type=%s",
            project_id, user.get("sub", "?"),
            bool(p.output_pptx), p.source_file_type,
        )

        if p.output_pptx:
            # Validate cached bytes before serving them
            ok, reason = validate_pptx_bytes(p.output_pptx)
            if ok:
                content = p.output_pptx
                log.info("Serving cached PPTX (%d bytes) for project %s", len(content), project_id)
            else:
                # Cached bytes are corrupt — rebuild and re-cache
                log.warning(
                    "Cached PPTX for project %s failed validation (%s); rebuilding",
                    project_id, reason,
                )
                p.output_pptx = None
                content = None
        else:
            content = None

        if content is None:
            if p.source_file_type == "pptx" and has_source_file(p):
                log.info(
                    "On-the-fly PPTX rebuild for project %s (source %d bytes, %d segments)",
                    project_id, len(get_source_bytes(p) or b""), len(p.segments),
                )
                _export_fidelity = "reconstructed"

                # Prefer the native PowerPoint COM backend (same as the
                # pipeline); fall back to the python-pptx rebuild on failure.
                from api.utils.powerpoint_com_finalizer import (
                    translate_pptx_with_powerpoint, powerpoint_com_available, PowerPointAutomationError,
                )
                if _native_office_enabled() and powerpoint_com_available():
                    try:
                        content = translate_pptx_with_powerpoint(get_source_bytes(p), p.segments, p.target_lang)
                        _export_fidelity = "native_office"
                    except Exception as _ppt_err:
                        log.warning(
                            "On-the-fly PowerPoint native translate failed (%s), falling back to reconstructed rebuild",
                            _ppt_err,
                        )
                        content = None

                if content is None:
                    try:
                        _layout_warnings: list[str] = []
                        content = rebuild_pptx(
                            get_source_bytes(p),
                            p.segments,
                            p.target_lang,
                            layout_warnings=_layout_warnings,
                            style_profile_override=_style_profile,
                            template_strength=_tpl_strength,
                            strict_qa=_strict_qa,
                            allow_export_with_warnings=_allow_export_with_warnings,
                            auto_repair_enabled=_auto_repair_enabled,
                            export_best_effort_result=_export_best_effort_result,
                        )
                        if _layout_warnings:
                            log.warning(
                                "PPTX export rebuild finished with unresolved layout warnings for %s: %s",
                                project_id,
                                " | ".join(_layout_warnings[:20]),
                            )
                    except Exception as exc:
                        log.error(
                            "PPTX rebuild failed for project %s: %s",
                            project_id, exc, exc_info=True,
                        )
                        raise HTTPException(
                            status_code=500,
                            detail=f"PPTX rebuild failed: {exc}",
                        )

                # Validate the freshly built bytes before caching and serving
                ok, reason = validate_pptx_bytes(content)
                if not ok:
                    log.error(
                        "Rebuilt PPTX for project %s is invalid (%s); size=%d",
                        project_id, reason, len(content),
                    )
                    raise HTTPException(
                        status_code=500,
                        detail=f"Rebuilt PPTX failed integrity check: {reason}",
                    )

                log.info(
                    "PPTX rebuild successful for project %s: %d bytes (fidelity=%s)",
                    project_id, len(content), _export_fidelity,
                )
                p.output_pptx = content
                p.formatting_fidelity = _export_fidelity
                db.commit()
            else:
                log.warning(
                    "No PPTX available for project %s: source_type=%s has_source_data=%s",
                    project_id, p.source_file_type, has_source_file(p),
                )
                raise HTTPException(400, "No PPTX output available for this project")

        dl_name_pptx = dl_filename if dl_filename.lower().endswith(".pptx") else (dl_filename.rsplit(".", 1)[0] + ".pptx" if "." in dl_filename else dl_filename + ".pptx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            headers={
                "Content-Disposition": content_disposition(dl_name_pptx),
                "Content-Length": str(len(content)),
                "Cache-Control": "private, no-store",
            },
        )

    elif fmt == "xlsx":
        if p.output_xlsx:
            content = p.output_xlsx
        elif p.source_file_type == "xlsx" and has_source_file(p):
            try:
                from api.utils.doc_rebuilder import rebuild_xlsx
                content = rebuild_xlsx(get_source_bytes(p), p.segments, p.target_lang)
                p.output_xlsx = content
                db.commit()
            except Exception as _e:
                log.warning("On-the-fly rebuild_xlsx failed (%s), cannot export XLSX", _e)
                raise HTTPException(500, f"XLSX rebuild failed: {_e}")
        else:
            raise HTTPException(400, "No XLSX output available for this project")
        dl_name_xlsx = dl_filename if dl_filename.lower().endswith(".xlsx") else (dl_filename.rsplit(".", 1)[0] + ".xlsx" if "." in dl_filename else dl_filename + ".xlsx")
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": content_disposition(dl_name_xlsx),
                "Content-Length": str(len(content)),
                "Cache-Control": "private, no-store",
            },
        )

    elif fmt == "txt":
        from api.utils.doc_rebuilder import build_translated_txt
        content = build_translated_txt(p.segments)
        dl_name_txt = dl_filename if dl_filename.lower().endswith(".txt") else (dl_filename.rsplit(".", 1)[0] + ".txt" if "." in dl_filename else dl_filename + ".txt")
        return Response(
            content=content,
            media_type="text/plain; charset=utf-8",
            headers={
                "Content-Disposition": content_disposition(dl_name_txt),
                "Content-Length": str(len(content)),
                "Cache-Control": "private, no-store",
            },
        )

    elif fmt == "html":
        from api.utils.doc_rebuilder import build_translated_html
        content = build_translated_html(p.segments, p.target_lang)
        dl_name_html = dl_filename if dl_filename.lower().endswith((".html", ".htm")) else (dl_filename.rsplit(".", 1)[0] + ".html" if "." in dl_filename else dl_filename + ".html")
        return Response(
            content=content,
            media_type="text/html; charset=utf-8",
            headers={
                "Content-Disposition": content_disposition(dl_name_html),
                "Content-Length": str(len(content)),
                "Cache-Control": "private, no-store",
            },
        )

    elif fmt == "pdf":
        # PDF output, best fidelity first:
        #  1) PDF source -> edit the ORIGINAL in place (keeps images/layout).
        #  2) DOCX->PDF via headless LibreOffice (matches the DOCX layout).
        #  3) ReportLab bilingual reflow (last-resort text-only PDF).
        pdf_bytes = None
        if p.source_file_type == "pdf" and has_source_file(p):
            try:
                from api.utils.pdf_inplace_translator import (
                    translate_pdf_in_place, InPlacePdfError,
                )
                pdf_bytes = translate_pdf_in_place(
                    get_source_bytes(p), p.segments, p.target_lang
                )
            except InPlacePdfError as _ip_err:
                log.info("In-place PDF not possible (%s); using reconstruction", _ip_err)
                pdf_bytes = None
            except Exception as _ip_err:
                log.warning("In-place PDF translation failed (%s); using reconstruction", _ip_err, exc_info=True)
                pdf_bytes = None
        if not pdf_bytes:
            try:
                from api.utils.libreoffice_finalizer import (
                    convert_docx_to_pdf_with_libreoffice, libreoffice_available,
                )
                if libreoffice_available():
                    if p.output_docx:
                        docx_for_pdf = p.output_docx  # already finalized by the pipeline
                    elif p.source_file_type == "docx" and has_source_file(p):
                        from api.utils.doc_rebuilder import rebuild_docx
                        docx_for_pdf = rebuild_docx(get_source_bytes(p), p.segments, p.target_lang)
                    else:
                        # PDF / TXT / HTML source — bilingual reconstructed DOCX.
                        from api.utils.doc_rebuilder import build_translated_docx
                        docx_for_pdf = build_translated_docx(p.segments, p.source_lang, p.target_lang)
                    pdf_bytes = convert_docx_to_pdf_with_libreoffice(docx_for_pdf)
            except Exception as _pdf_conv_err:
                log.warning(
                    "DOCX->PDF via LibreOffice failed (%s); falling back to ReportLab PDF",
                    _pdf_conv_err,
                )
                pdf_bytes = None
        if not pdf_bytes:
            from api.utils.doc_rebuilder import build_translated_pdf
            pdf_bytes = build_translated_pdf(
                p.segments, p.source_lang, p.target_lang, project_name=p.name
            )
        dl_name_pdf = dl_filename if dl_filename.lower().endswith(".pdf") else (dl_filename.rsplit(".", 1)[0] + ".pdf" if "." in dl_filename else dl_filename + ".pdf")
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": content_disposition(dl_name_pdf),
                "Content-Length": str(len(pdf_bytes)),
                "Cache-Control": "private, no-store",
            },
        )

    else:
        raise HTTPException(400, f"Unsupported export format: {fmt}")


# ── Custom Dictionary endpoints ───────────────────────────────────────────────

@router.get("/translation/dictionary")
def list_dictionary(
    source_lang: str = "en",
    target_lang: str = "ar",
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    uid = _user_id(user)
    # Own entries + shared (user_id=NULL); always scoped, never unfiltered
    entries = (
        db.query(CustomDictionaryEntry)
        .filter(
            CustomDictionaryEntry.source_lang == source_lang,
            CustomDictionaryEntry.target_lang == target_lang,
            (CustomDictionaryEntry.user_id == uid) | (CustomDictionaryEntry.user_id.is_(None)),
        )
        .order_by(CustomDictionaryEntry.source_term)
        .all()
    )
    return [
        {
            "id": e.id,
            "source_term": e.source_term,
            "target_term": e.target_term,
            "source_lang": e.source_lang,
            "target_lang": e.target_lang,
            "domain": e.domain,
            "notes": e.notes,
            "is_shared": e.user_id is None,
            "created_at": e.created_at.isoformat() if e.created_at else None,
        }
        for e in entries
    ]


class DictEntryBody(BaseModel):
    source_term: str
    target_term: str
    source_lang: str = "en"
    target_lang: str = "ar"
    domain: Optional[str] = None
    notes: Optional[str] = None


@router.post("/translation/dictionary")
def add_dictionary_entry(
    body: DictEntryBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    entry = CustomDictionaryEntry(
        user_id=user["id"],
        source_term=body.source_term.strip(),
        target_term=body.target_term.strip(),
        source_lang=body.source_lang,
        target_lang=body.target_lang,
        domain=body.domain,
        notes=body.notes,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    return {"id": entry.id, "source_term": entry.source_term, "target_term": entry.target_term}


@router.put("/translation/dictionary/{entry_id}")
def update_dictionary_entry(
    entry_id: str,
    body: DictEntryBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    entry = _get_owned_dict_entry(db, entry_id, user)
    entry.source_term = body.source_term.strip()
    entry.target_term = body.target_term.strip()
    entry.source_lang = body.source_lang
    entry.target_lang = body.target_lang
    entry.domain = body.domain
    entry.notes = body.notes
    db.commit()
    return {"id": entry.id, "source_term": entry.source_term, "target_term": entry.target_term}


@router.delete("/translation/dictionary/{entry_id}")
def delete_dictionary_entry(
    entry_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    entry = _get_owned_dict_entry(db, entry_id, user)
    db.delete(entry)
    db.commit()
    return {"deleted": entry_id}


@router.post("/translation/dictionary/{entry_id}/share")
def share_dictionary_entry(
    entry_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """
    Promote a personal dictionary entry to a team-shared entry (user_id = NULL).
    Any team member will then see and use this entry during translation.
    Only the owner may share their own entry.
    """
    entry = _get_owned_dict_entry(db, entry_id, user)
    entry.user_id = None
    db.commit()
    return {"shared": entry_id, "source_term": entry.source_term}


class ShareAllBody(BaseModel):
    source_lang: str = "en"
    target_lang: str = "ar"


@router.post("/translation/dictionary/share-all")
def share_all_dictionary_entries(
    body: ShareAllBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """
    Promote ALL of the caller's personal dictionary entries (for a language pair)
    to shared entries (user_id = NULL).
    """
    uid = _user_id(user)
    entries = (
        db.query(CustomDictionaryEntry)
        .filter(
            CustomDictionaryEntry.user_id == uid,
            CustomDictionaryEntry.source_lang == body.source_lang,
            CustomDictionaryEntry.target_lang == body.target_lang,
        )
        .all()
    )
    for e in entries:
        e.user_id = None
    db.commit()
    return {"shared_count": len(entries)}


@router.post("/translation/dictionary/import")
async def import_dictionary(
    file: UploadFile = File(...),
    source_lang: str = Form("en"),
    target_lang: str = Form("ar"),
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Import terminology from a TSV file (source_term TAB target_term [TAB domain])."""
    import csv as csv_mod
    content = await file.read()
    text = content.decode("utf-8", errors="replace")
    added = 0
    reader = csv_mod.reader(io.StringIO(text), delimiter="\t")
    for row in reader:
        if len(row) < 2:
            continue
        source_term = row[0].strip()
        target_term = row[1].strip()
        domain = row[2].strip() if len(row) > 2 else None
        if source_term and target_term:
            db.add(CustomDictionaryEntry(
                user_id=user["id"],
                source_term=source_term,
                target_term=target_term,
                source_lang=source_lang,
                target_lang=target_lang,
                domain=domain,
            ))
            added += 1
    db.commit()
    return {"imported": added}


@router.get("/translation/dictionary/export")
def export_dictionary(
    source_lang: str = "en",
    target_lang: str = "ar",
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """Export custom dictionary as TSV."""
    import csv as csv_mod
    uid = _user_id(user)
    # Own entries + shared (user_id=NULL); always scoped, never unfiltered
    entries = (
        db.query(CustomDictionaryEntry)
        .filter(
            CustomDictionaryEntry.source_lang == source_lang,
            CustomDictionaryEntry.target_lang == target_lang,
            (CustomDictionaryEntry.user_id == uid) | (CustomDictionaryEntry.user_id.is_(None)),
        )
        .all()
    )

    buf = io.StringIO()
    writer = csv_mod.writer(buf, delimiter="\t")
    writer.writerow(["source_term", "target_term", "domain", "notes"])
    for e in entries:
        writer.writerow([e.source_term, e.target_term, e.domain or "", e.notes or ""])

    return Response(
        content=buf.getvalue().encode("utf-8"),
        media_type="text/tab-separated-values; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="dictionary.tsv"'},
    )


# ── Translation Memory endpoints ──────────────────────────────────────────────

@router.get("/translation/memory")
def list_memory(
    source_lang: str = "en",
    target_lang: str = "ar",
    search: str = "",
    limit: int = 100,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    uid = _user_id(user)
    # Own entries + shared (user_id=NULL); always scoped, never unfiltered
    q = db.query(TranslationSegment).filter(
        TranslationSegment.source_lang == source_lang,
        TranslationSegment.target_lang == target_lang,
        (TranslationSegment.user_id == uid) | (TranslationSegment.user_id.is_(None)),
    )
    if search:
        q = q.filter(
            TranslationSegment.source_text.ilike(f"%{search}%") |
            TranslationSegment.target_text.ilike(f"%{search}%")
        )
    entries = q.order_by(TranslationSegment.use_count.desc()).limit(limit).all()
    return [
        {
            "id": e.id,
            "source_text": e.source_text,
            "target_text": e.target_text,
            "use_count": e.use_count,
            "is_shared": e.user_id is None,
            "updated_at": e.updated_at.isoformat() if e.updated_at else None,
        }
        for e in entries
    ]


class MemoryEditBody(BaseModel):
    target_text: str


@router.put("/translation/memory/{memory_id}")
def update_memory(
    memory_id: str,
    body: MemoryEditBody,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    entry = _get_owned_memory_entry(db, memory_id, user)
    entry.target_text = body.target_text
    db.commit()
    return {"id": entry.id, "target_text": entry.target_text}


@router.delete("/translation/memory/{memory_id}")
def delete_memory(
    memory_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    entry = _get_owned_memory_entry(db, memory_id, user)
    db.delete(entry)
    db.commit()
    return {"deleted": memory_id}


@router.post("/translation/memory/{memory_id}/share")
def share_memory_entry(
    memory_id: str,
    db: Session = Depends(get_db),
    user: dict = Depends(require_auth),
):
    """
    Promote a personal translation memory entry to a team-shared entry (user_id = NULL).
    All team members will benefit from this cached translation on future projects.
    Only the owner may share their own entry.
    """
    entry = _get_owned_memory_entry(db, memory_id, user)
    entry.user_id = None
    db.commit()
    return {"shared": memory_id}
